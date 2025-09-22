import os
import subprocess
from pathlib import Path
from typing import List

import pandas as pd
from loguru import logger
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PyPDF2 import PdfMerger

from .config import Config
from .invoice_context import InvoiceContext


class DocumentUtils:
    """
    Statische Hilfsklasse für Dokumentenoperationen:
    - DOCX nach PDF konvertieren
    - PDFs zusammenführen
    - Rechnungsübersicht als Excel-Tabelle erstellen

    """

    @staticmethod
    def docx_to_pdf(docx_path: Path, pdf_path: Path, invoice_context: InvoiceContext) -> Path:
        """
        Konvertiert eine DOCX-Datei in eine PDF-Datei mit LibreOffice und benennt sie ggf. nach Vorgabe um.

        Args:
            docx_path (Path): Pfad zur DOCX-Datei.
            pdf_path (Path): Pfad zur Ausgabedatei (PDF).
            invoice_context (InvoiceContext): Kontextobjekt mit allen Rechnungsdaten.
        Returns:
            Path: Pfad zur erzeugten PDF-Datei.
        Raises:
            RuntimeError: Wenn die Konvertierung oder Umbenennung fehlschlägt.
        """
        try:
            subprocess.run(
                [
                    "libreoffice",
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(pdf_path.parent),
                    str(docx_path),
                ],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=True,
            )
        except Exception as e:
            logger.error(f"PDF-Konvertierung fehlgeschlagen: {e}")
            raise RuntimeError(f"PDF-Konvertierung fehlgeschlagen: {e}")

        generated_pdf = docx_path.with_suffix(".pdf")
        # Dateinamen mit Entitäten und Leistungszeitraum
        payer_id = invoice_context.data.get("payer").key
        client_id = invoice_context.data.get("client").key
        invoice_month = invoice_context.data.get("invoice_month", "unbekannt")
        target_name = f"Rechnung_{payer_id}_{client_id}_{invoice_month}"
        target_pdf = pdf_path.parent / target_name
        try:
            os.rename(generated_pdf, target_pdf)
        except Exception as e:
            logger.error(f"PDF konnte nicht umbenannt werden: {e}")
            raise RuntimeError(f"PDF konnte nicht umbenannt werden: {e}")
        logger.debug(f"{target_pdf.name} erzeugt")
        return target_pdf

    @staticmethod
    def merge_pdfs(
        pdf_files: List[Path], payer_context: InvoiceContext, output_path: Path = None
    ) -> Path:
        """
        Führt mehrere PDF-Dateien zu einer einzigen zusammen.

        Args:
            pdf_files (List[Path]): Liste der PDF-Dateipfade.
            invoice_context (InvoiceContext): Kontextobjekt mit Zahlungsdienstleister und Zeitraum.
            output_path (Path, optional): Zielverzeichnis für die Sammel-PDF.
        Returns:
            Path: Pfad zur erzeugten PDF-Datei.
        Raises:
            ValueError: Wenn keine PDF-Dateien übergeben wurden.
            FileNotFoundError: Wenn eine der PDF-Dateien nicht gefunden wird.
            RuntimeError: Wenn die Zusammenführung fehlschlägt.
        """
        if not pdf_files:
            raise ValueError("Keine PDF-Dateien zum Zusammenführen gefunden.")
        merger = PdfMerger()
        for pdf_file in pdf_files:
            if not pdf_file.exists():
                raise FileNotFoundError(f"PDF-Datei nicht gefunden: {pdf_file}")
            merger.append(pdf_file)
        # Zielverzeichnis bestimmen
        if output_path is None:
            output_path = pdf_files[0].parent
        payer = payer_context.data.get("payer")
        invoice_month = payer_context.data.get("invoice_month", "unbekannt")
        merged_pdf_path = output_path / f"Rechnungen_{payer.key}_{invoice_month}.pdf"
        try:
            merger.write(merged_pdf_path)
            merger.close()
        except Exception as e:
            logger.error(f"PDF-Zusammenführung fehlgeschlagen: {e}")
            raise RuntimeError(f"PDF-Zusammenführung fehlgeschlagen: {e}")
        return merged_pdf_path

    @staticmethod
    def create_summary(
        config: Config,
        invoice_list: List[InvoiceContext],
    ) -> Path:
        """
        Erstellt eine Excel-Datei mit der Rechnungsübersicht.
        Es werden nur die Felder invoice_date, payer_id, client_id, invoice_month und Summe_Kosten aufgenommen.
        Args:
            config (Config): Konfigurationsobjekt für Pfade und Formate.
            invoice_list (List[InvoiceContext]): Liste der Rechnungskontexte.
        Returns:
            Path: Pfad zur erzeugten Excel-Datei.
        Raises:
            RuntimeError: Wenn die Datei nicht geschrieben werden kann.
        """
        output_path: Path = Path(config.data["structure"]["output_path"])
        kosten_format = config.data.get("currency_format", '#,##0.00 "CHF"')
        datum_format = config.data.get("date_format", "DD.MM.YY")

        # Nur die gewünschten Felder extrahieren
        summary_rows = []
        for invoice in invoice_list:
            data = invoice.as_dict()
            summary_rows.append(
                {
                    "Rechnungsdatum": data.get("invoice_date"),
                    "ZD-Nr": getattr(data.get("payer"), "key", "n.a"),
                    "Klienten-Nr": getattr(data.get("client"), "key", "n.a"),
                    "Abrechnungsmonat": data.get("invoice_month", "unbekannt"),
                    "Rechnungsbetrag": data.get("summe_kosten", -999),
                }
            )
        summary_df = pd.DataFrame(summary_rows)
        # Prüfe, ob alle Rechnungen denselben Abrechnungsmonat haben und propagiere diesen
        invoice_months = set(row["Abrechnungsmonat"] for row in summary_rows)
        if len(invoice_months) == 1:
            invoice_month = invoice_months.pop()
        else:
            logger.warning(
                f"Mehrere Abrechnungsmonate in der Übersicht: {invoice_months}"
            )
            invoice_month = "gemischt"

        summary_file = output_path / f"Rechnungsuebersicht_{invoice_month}.xlsx"
        try:
            with pd.ExcelWriter(summary_file, engine="openpyxl") as writer:
                summary_df.to_excel(
                    writer, index=False, sheet_name="Rechnungsübersicht"
                )
                worksheet = writer.sheets["Rechnungsübersicht"]
                end_col_idx = len(summary_df.columns)
                end_col_letter = get_column_letter(end_col_idx)
                table_ref = f"A1:{end_col_letter}{len(summary_df) + 1}"
                tab = Table(displayName="Rechnungsübersicht", ref=table_ref)
                tab.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                worksheet.add_table(tab)
                # Formatierung und Summen für die Rechnungsbetrag-Spalte
                if "Rechnungsbetrag" in summary_df.columns:
                    kosten_col_idx = (
                        int(summary_df.columns.get_loc("Rechnungsbetrag")) + 1
                    )
                    kosten_col_letter = get_column_letter(kosten_col_idx)
                    # Format aus der Config holen
                    for row in range(2, len(summary_df) + 2):
                        worksheet[
                            f"{kosten_col_letter}{row}"
                        ].number_format = kosten_format
                    total_row_idx = len(summary_df) + 2
                    worksheet[f"A{total_row_idx}"] = "Gesamt"
                    worksheet[f"{kosten_col_letter}{total_row_idx}"] = (
                        f"=SUM({kosten_col_letter}2:{kosten_col_letter}{total_row_idx - 1})"
                    )
                    worksheet[
                        f"{kosten_col_letter}{total_row_idx}"
                    ].number_format = kosten_format
                    bold_font = Font(bold=True)
                    for col in range(1, end_col_idx + 1):
                        worksheet.cell(row=total_row_idx, column=col).font = bold_font
                if "Rechnungsdatum" in summary_df.columns:
                    datum_col_idx = (
                        int(summary_df.columns.get_loc("Rechnungsdatum")) + 1
                    )
                    datum_col_letter = get_column_letter(datum_col_idx)
                    for row in range(2, len(summary_df) + 2):
                        worksheet[
                            f"{datum_col_letter}{row}"
                        ].number_format = datum_format
        except Exception as e:
            logger.error(f"Fehler beim Schreiben der Excel-Datei: {e}")
            raise RuntimeError(f"Fehler beim Schreiben der Excel-Datei: {e}")
        return summary_file


if __name__ == "__main__":
    print("DocumentUtils Modul. Nicht direkt ausführbar.")
