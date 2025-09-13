import subprocess
import tempfile
from io import BytesIO
from pathlib import Path
from typing import Optional, Tuple
from zipfile import ZipFile

import pandas as pd
import qrcode
import yaml
from docx.shared import Mm
from docxtpl import DocxTemplate, InlineImage  # type: ignore
from icecream import ic  # type: ignore
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageDraw, ImageFont
from PyPDF2 import PdfMerger  # type: ignore
from rich import print
from rich.traceback import install

# Globals
CONFIG: dict = {}


def load_config():
    global CONFIG
    cwd = Path(__file__).parent.parent.parent / ".config"
    with open(cwd / "wegpiraten_config.yaml", "r") as f:
        CONFIG = yaml.safe_load(f)


def format_2f(value: float, currency: Optional[str] = None) -> str:
    if pd.isna(value):
        return ""
    currency = currency or ""
    if currency and not currency.startswith(" "):
        currency = " " + currency
    # Tausenderpunkt, Dezimalkomma
    tmp_val = f"{value:,.2f}"
    tmp_val = tmp_val.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{tmp_val}{currency}"


def clear_path(path: Path):
    """Löscht alle Dateien im angegebenen Verzeichnis

    Args:
        path (Path): Verzeichnis, dessen Dateien gelöscht werden sollen
    """
    for item in path.iterdir():
        if item.is_file():
            item.unlink()


def zip_docs(src_dir: Path, zip_path: Path):
    with ZipFile(zip_path, "w") as zipf:
        for file in src_dir.glob("Rechnung_*.docx"):
            zipf.write(file, arcname=file.name)


def load_data(
    db: Path,
    sheet: Optional[str] = None,
    abrechnungsmonat: Optional[pd.Timestamp | str] = None,
) -> pd.DataFrame:
    """
    Lade die Aufwandsdaten aus der Excel-DB
    Achtung: Filter in der Excel-Abfrage bleiben bestehen

    Args:
        db (Path): Ort der Datenbank (Excel-Datei)
        sheet (str): Name des Arbeitsblatts. Wenn None, wird das aktive Blatt verwendet.
        abrechnungsmonat (pd.Timestamp): Abrechnungsmonat. Wenn None, wird

    Returns:
        pd.DataFrame: DataFrame mit den Aufwandsdaten für den Abrechnungsmonat
    """
    work_book = load_workbook(db, data_only=True)
    work_sheet = work_book[sheet] if sheet else work_book.active

    # Monat als pd.Timestamp
    if abrechnungsmonat is None:
        abrechnungsmonat = pd.Timestamp.now().to_period("M").to_timestamp()
    elif isinstance(abrechnungsmonat, str):
        abrechnungsmonat = pd.to_datetime(abrechnungsmonat, format="%Y-%m")
    elif not isinstance(abrechnungsmonat, pd.Timestamp):
        raise ValueError(
            "abrechnungsmonat muss ein String im Format 'YYYY-MM' oder ein pd.Timestamp sein"
        )

    # Generator für die Zeilen
    data = work_sheet.values

    # Erste drei Zeilen sind Metadaten
    # Eventuell besser, bis zu einem Schlüsselwort zu springen
    for _ in range(3):
        next(data)

    # Erste Zeile sind die Spaltennamen, aber ohne die erste Spalte (ID)
    columns = next(data)[1:]

    # Daten ab zweiter Spalte
    df = pd.DataFrame(
        (row[1:] for row in data),
        columns=columns,
    )

    # df["Leistungsdatum"] = pd.to_datetime(
    #     df["Leistungsdatum"], errors="coerce", format="%d.%m.%Y"
    # )

    # Start und Ende des Monats bestimmen
    monat_start: pd.Timestamp = abrechnungsmonat
    monat_ende: pd.Timestamp = abrechnungsmonat + pd.offsets.MonthEnd(0)

    # Nur Leistungen im Abrechnungsmonat übernehmen
    df = df[
        (df["Leistungsdatum"] >= monat_start) & (df["Leistungsdatum"] <= monat_ende)
    ]

    # Fehlende Werte in ZD_Name2 mit Leerzeichen auffüllen/ersetzen
    df["ZD_Name2"] = df["ZD_Name2"].fillna("").replace("(Leer)", "")

    return df


def check_data_consistency(df: pd.DataFrame):
    """Überprüfe, ob alle erwarteten Spalten im DataFrame vorhanden sind

    Args:
        df (pd.DataFrame): DataFrame mit den Aufwandsdaten
        expected_columns (list): Liste der erwarteten Spaltennamen

    Raises:
        ValueError: Wenn eine oder mehrere erwartete Spalten fehlen
    """

    expected_columns = {
        col["name"] if isinstance(col, dict) else col
        for col in CONFIG["expected_columns"]
    }

    missing_columns = expected_columns - set(df.columns)

    if missing_columns:
        missing_str = "\n".join(sorted(missing_columns))
        print(f"Fehlende Spalten: {missing_str}")
        raise ValueError(f"Fehlende Felder in der Pivot-Tabelle: {missing_str}")


def create_invoice_id(
    inv_month: Optional[pd.Timestamp | str] = None, client_id: Optional[str] = None
) -> str:
    """Rechnungsnummer aus AbrMon und Klienten-ID generieren

    Args:
        inv_month (pd.Timestamp): Abrechnungsmonat
        client_id (str): Klienten-ID

    Returns:
        str: Rechnungsnummer als String
    """
    #
    # TODO: Klären, ob die erzeugte ReNr ok ist

    if inv_month is None:
        inv_month = pd.Timestamp.now()
    if isinstance(inv_month, str):
        # Extract year and month, format as MMYY
        dt = pd.to_datetime(
            inv_month.replace(".", "-"), format="%m-%Y", errors="coerce"
        )
        if dt is pd.NaT:
            raise ValueError(
                "inv_month muss ein String im Format 'MM-YYYY' oder ein pd.Timestamp sein"
            )
        month_mmYY = dt.strftime("%m%y")
    elif isinstance(inv_month, pd.Timestamp):
        month_mmYY = inv_month.strftime("%m%y")
    else:
        raise ValueError(
            "inv_month muss ein String im Format 'YYYY-MM' oder ein pd.Timestamp sein"
        )
    client_id = client_id or "K000"

    return f"R{month_mmYY}_{client_id}"


def create_einzahlungsschein_png_dynamic(
    context: dict,
    output_png: str,
    font_dir: str = "/usr/share/fonts/truetype/msttcorefonts/",
):
    """
    Erstelle einen Einzahlungsschein als PNG-Datei mit dynamischen Daten
    Args:
        context (dict): Dictionary mit den dynamischen Daten für den Einzahlungsschein
        output_png (str): Pfad zur Ausgabedatei (PNG)   
        font_dir (str): [Optional] Verzeichnis, in dem die Calibri-Fonts liegen    
    """
    width, height = 1800, 900
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)

    # Calibri-Font laden (Pfad ggf. anpassen!)
    try:
        font_path = Path(font_dir) / "calibri.ttf"
        font_path_bold = Path(font_dir) / "calibrib.ttf"
        font = ImageFont.truetype(font_path, 36)
        font_bold = ImageFont.truetype(font_path_bold, 48)
        font_small = ImageFont.truetype(font_path, 28)
    except Exception as e:
        print(
            "[red on white]Warnung: Calibri-Font nicht gefunden, Standardfont wird verwendet.",
            e,
        )
        font = font_bold = font_small = ImageFont.load_default()

    # Empfängerdaten aus CONFIG holen und direkt verwenden
    empf_IBAN: str = CONFIG["empfaenger"]["IBAN"]
    empf_name: str = CONFIG["empfaenger"]["name"]
    empf_strasse: str = CONFIG["empfaenger"]["strasse"]
    empf_adresse: str = CONFIG["empfaenger"]["plz_ort"]

    # Linien
    draw.line([(width // 2, 60), (width // 2, height - 60)], fill="black", width=3)
    draw.line([(60, 60), (width - 60, 60)], fill="black", width=2)

    # --- Linker Bereich: Empfangsschein ---
    x1, y1 = 80, 100
    draw.text((x1, y1), "Empfangsschein", font=font_bold, fill="black")
    y1 += 60

    draw.text((x1, y1), "Konto / Zahlbar an", font=font_small, fill="black")
    y1 += 40
    draw.text((x1, y1), empf_IBAN, font=font, fill="black")
    y1 += 40
    draw.text((x1, y1), empf_name, font=font, fill="black")
    y1 += 40
    draw.text((x1, y1), empf_strasse, font=font, fill="black")
    y1 += 60

    draw.text((x1, y1), "Zahlbar durch", font=font_small, fill="black")
    y1 += 40
    draw.text((x1, y1), context["ZD_Name"], font=font, fill="black")
    y1 += 40
    draw.text((x1, y1), context["ZD_Strasse"], font=font, fill="black")
    y1 += 60

    draw.text((x1, y1), "Währung", font=font_small, fill="black")
    draw.text((x1 + 180, y1), "Betrag", font=font_small, fill="black")
    y1 += 40
    draw.text((x1, y1), "CHF", font=font, fill="black")
    draw.text((x1 + 180, y1), context["Summe_Kosten_2f"], font=font, fill="black")

    # --- Rechter Bereich: Zahlteil ---
    x2, y2 = width // 2 + 80, 100
    draw.text((x2, y2), "Zahlteil", font=font_bold, fill="black")
    y2 += 60

    draw.text((x2, y2), "Konto / Zahlbar an", font=font_small, fill="black")
    y2 += 40
    draw.text((x2, y2), empf_IBAN, font=font, fill="black")
    y2 += 40
    draw.text((x2, y2), empf_name, font=font, fill="black")
    y2 += 40
    draw.text((x2, y2), empf_strasse, font=font, fill="black")
    y2 += 40

    draw.text((x2, y2), "Zusätzliche Informationen", font=font_small, fill="black")
    y2 += 40
    draw.text((x2, y2), context["Rechnungsnummer"], font=font, fill="black")
    y2 += 40

    draw.text((x2, y2), "Zahlbar durch", font=font_small, fill="black")
    y2 += 40
    draw.text((x2, y2), context["ZD_Name"], font=font, fill="black")
    y2 += 40
    draw.text((x2, y2), context["ZD_Strasse"], font=font, fill="black")
    y2 += 60

    draw.text((x2, y2), "Währung", font=font_small, fill="black")
    draw.text((x2 + 180, y2), "Betrag", font=font_small, fill="black")
    y2 += 40
    draw.text((x2, y2), "CHF", font=font, fill="black")
    draw.text((x2 + 180, y2), context["Summe_Kosten_2f"], font=font, fill="black")

    # --- QR-Code generieren und einfügen ---
    qr_data = f"""SPC
0200
1
{empf_IBAN}
{empf_name}
{empf_strasse}
{context["Summe_Kosten_2f"]}
CHF
{context["Rechnungsnummer"]}
{context["ZD_Name"]}
{context["ZD_Strasse"]}
"""
    qr = qrcode.make(qr_data)
    qr = qr.resize((300, 300))
    img.paste(qr, (width - 350, height // 2 - 150))

    img.save(output_png)


def format_fields(df: pd.DataFrame) -> pd.DataFrame:
    """
    Formatiere die Felder im DataFrame entsprechend den Vorgaben in expected_columns
    (z.B. Zahlen mit 2 Nachkommastellen, CHF-Suffix etc.)

    Args:
        df (pd.DataFrame): DataFrame mit den Aufwandsdaten
        expected_columns (list): Liste der erwarteten Spaltennamen und deren Formate

    Returns:
        pd.DataFrame: DataFrame mit formatierten Feldern
    """
    num_fields = []
    # Liste der numerischen Felder und optional die Währung aus den CONFIG holen
    expected_columns = CONFIG["expected_columns"]
    for col in expected_columns:
        if isinstance(col, dict) and col.get("type") in ("numeric", "currency"):
            currency = col.get("currency")
            num_fields.append((col["name"], currency))

    # Neue Spalten mit _2f-Suffix für die formatierten Werte erstellen
    for col, currency in num_fields:
        decimals = next(
            (
                c.get("decimals", 2)
                for c in expected_columns
                if isinstance(c, dict) and c["name"] == col
            ),
            2,
        )
        df[f"{col}_2f"] = df[col].apply(
            lambda x: format_2f(x, currency) if pd.notna(x) else ""
        )
    date_format = next(
        (
            c.get("format")
            for c in expected_columns
            if isinstance(c, dict) and c["name"] == "Leistungsdatum"
        ),
        "%d.%m.%Y",
    )
    df["Leistungsdatum"] = df["Leistungsdatum"].dt.strftime(date_format)

    return df


def format_invoice(
    client_details: pd.DataFrame,
) -> Tuple[DocxTemplate, pd.DataFrame]:
    """
    Erstelle eine Rechnung für einen Klienten
    Zur Übernahme in das Template müssen viele Werte formatiert
    (z.B. Zahlen mit 2 Nachkommastellen, CHF-Suffix etc.)
    oder neu erstellt werden (Summen).

    Args:
        invoice_template (DocxTemplate): Vorlage für die Rechnung
        invoice_id (str): Rechnungsnummer
        client_id (str): Klienten-ID
        client_details (pd.DataFrame): Details für diesen Klienten, config Kontext
    Returns:
        DocxTemplate: Ausgefüllte Rechnung
    """

    invoice_template = DocxTemplate((Path(CONFIG["structure"]["prj_root"]) / 
                                     CONFIG["structure"]["template_path"] / 
                                     CONFIG["invoice_template_name"])
    )

    # Summen über relevante Spalten in die Daten übernehmen
    # client_details ist eine Gruppe aus dem DataFrame
    # (alle Zeilen für einen Klienten) + Empfängerdaten
    client_details["Summe_Fahrtzeit"] = client_details["Fahrtzeit"].sum()
    client_details["Summe_Direkt"] = client_details["Direkt"].sum()
    client_details["Summe_Indirekt"] = client_details["Indirekt"].sum()
    client_details["Summe_Stunden"] = client_details["Stunden"].sum()
    client_details["Summe_Kosten"] = client_details["Kosten"].sum()
    # Diese Felder dem dict CONFIG["expected_columns"] hinzufügen
    client_details["Summe_Fahrtzeit_2f"] = format_2f(
        client_details["Summe_Fahrtzeit"].iloc[0]
    )
    client_details["Summe_Direkt_2f"] = format_2f(
        client_details["Summe_Direkt"].iloc[0]
    )
    client_details["Summe_Indirekt_2f"] = format_2f(
        client_details["Summe_Indirekt"].iloc[0]
    )
    client_details["Summe_Stunden_2f"] = format_2f(
        client_details["Summe_Stunden"].iloc[0]
    )
    client_details["Summe_Kosten_2f"] = format_2f(
        client_details["Summe_Kosten"].iloc[0], "CHF"
    )

    # Liste der sonstigen numerischen Felder und optional die Währung
    client_details = format_fields(client_details)

    # Datumsfelder als Strings ohne Uhrzeit formatieren
    # Datumsformat aus der Config holen

    # Abrechnungsmonat nur als String im Format 'mm-YYYY'. Das geht, weil die Leistungsdaten
    # nur für einen Monat vorliegen (Filter in load_data)
    abrechnungsmonat: str = client_details["Leistungsdatum"].iloc[0][3:]
    client_details["Abrechnungsmonat"] = abrechnungsmonat
    client_details["Rechnungsdatum"] = pd.Timestamp.now().strftime("%d.%m.%Y")

    # Rechnungsnummer generieren
    invoice_id = create_invoice_id(
        inv_month=abrechnungsmonat,
        client_id=client_details["Klient-Nr."].iloc[0],
    )

    client_details["Rechnungsnummer"] = invoice_id

    print(f"Erstelle Rechnung {invoice_id}")

    # Kopf-Daten (sind für alle Zeilen der Gruppe gleich,
    # nehmen wir die aus der ersten Zeile (Index 0))
    # Summen üer relevante Spalten in die Daten übernehmen
    invoice_context = client_details.iloc[0].to_dict()

    # Tabellen-Daten (alle Zeilen für diese Gruppe)
    invoice_positions = client_details[
        [
            "Leistungsdatum",
            "Fahrtzeit_2f",
            "Direkt_2f",
            "Indirekt_2f",
            "Sollstunden_2f",
            # "Stundensatz_2f",
            "km_Pauschale_2f",
            "Stunden_2f",
            "Kosten_2f",
        ]
    ]

    invoice_details: dict = invoice_positions.to_dict(orient="records")

    # Einzahlungsschein als PNG erstellen
    output_png = CONFIG["structure"]["tmp_path"]  # Temporärer Pfad

    # Erzeuge den Einzahlungsschein als temporäre Datei und lies diese gleich wieder ein
    with tempfile.NamedTemporaryFile(
        dir=output_png, suffix=".png", delete=True
    ) as tmp_file:
        create_einzahlungsschein_png_dynamic(invoice_context, tmp_file.name)
        # Einzahlungsschein in die Rechnung einfügen
        einzahlungsschein_img = InlineImage(
            invoice_template, tmp_file.name, width=Mm(200)
        )

        # Kontext fürs Template
        # Felder im Template müssen mit den Keys im Kontext
        # exakt übereinstimmen (inkl. Groß-/Kleinschreibung)
        context = {
            **invoice_context,
            "Positionen": invoice_details,
            "Einzahlungsschein": einzahlungsschein_img,
        }

        # Rendern
        invoice_template.render(context)
    return invoice_template, client_details


def docx_to_pdf(docx_path, pdf_path):
    """
    Konvertiere eine DOCX-Datei in eine PDF-Datei mit LibreOffice

    Args:
        docx_path (str): Pfad zur DOCX-Datei
        pdf_path (str): Pfad zur Ausgabedatei (PDF)
    """

    subprocess.run(
        [
            "libreoffice",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(pdf_path.parent),
            str(docx_path),
        ],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        check=True,
    )


def merge_pdfs(pdf_files: list, zdnr: str, abrechnungsmonat: str):
    """
    Merge multiple PDF files into a single PDF.

    Args:
        pdf_files (list): List of absolute paths to PDF files to be merged.
        zdnr (str): Identifier for the payment service provider.
            abrechnungsmonat (str): Billing month in 'YYYY-MM' format.
    """

    merger = PdfMerger()

    # Use the directory of the first PDF file as the output path
    output_path = pdf_files[0].parent

    for pdf_file in pdf_files:
        merger.append(pdf_file)

    merged_pdf_path = output_path / f"Rechnungen_{zdnr}_{abrechnungsmonat}.pdf"
    merger.write(merged_pdf_path)
    merger.close()

    ic("Zusammengefasste PDF gespeichert:", str(merged_pdf_path.name))


def create_summary(summary_rows: list, abrechnungsmonat: str):
    """Erstelle eine Excel-Datei mit der Rechnungsübersicht

    Args:
        summary_rows (list): Liste der Rechnungsübersicht-Zeilen
        output_path (Path): Pfad zum Speichern der Excel-Datei
        abrechnungsmonat (str): Abrechnungsmonat im Format 'YYYY-MM'
    """
    output_path: Path = Path(CONFIG["structure"]["output_path"])

    summary_df = pd.DataFrame(summary_rows)
    summary_file = output_path / f"Rechnungsuebersicht_{abrechnungsmonat}.xlsx"

    with pd.ExcelWriter(summary_file, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Rechnungsübersicht")
        worksheet = writer.sheets["Rechnungsübersicht"]

        # Endspalte dynamisch bestimmen
        end_col_idx = len(summary_df.columns)
        end_col_letter = chr(64 + end_col_idx)  # 64 + 1 = 'A', etc.
        table_ref = f"A1:{end_col_letter}{len(summary_df) + 1}"

        tab = Table(displayName="Rechnungsübersicht", ref=table_ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(tab)

        # Spalte "Summe_Kosten" als Währung formatieren
        if "Summe_Kosten" in summary_df.columns:
            kosten_col_idx = summary_df.columns.get_loc("Summe_Kosten") + 1
            kosten_col_letter = chr(64 + kosten_col_idx)
            for row in range(2, len(summary_df) + 2):
                worksheet[f"{kosten_col_letter}{row}"].number_format = '#,##0.00 "CHF"'

            # Ergebniszeile hinzufügen
            total_row_idx = len(summary_df) + 2
            worksheet[f"A{total_row_idx}"] = "Gesamt"
            worksheet[f"{kosten_col_letter}{total_row_idx}"] = (
                f"=SUM({kosten_col_letter}2:{kosten_col_letter}{total_row_idx - 1})"
            )
            worksheet[
                f"{kosten_col_letter}{total_row_idx}"
            ].number_format = '#,##0.00 "CHF"'
            # Fettdruck für die Ergebniszeile
            bold_font = Font(bold=True)
            for col in range(1, end_col_idx + 1):
                worksheet.cell(row=total_row_idx, column=col).font = bold_font

    ic("Rechnungsübersicht gespeichert:" + str(summary_file))


def main():
    # Rich Traceback für bessere Fehlermeldungen
    install(show_locals=True)

    # Alle Konfigurationsdaten laden
    load_config()
    env: dict = CONFIG["structure"]
    project_root = Path(env["prj_root"])
    tmp_path: Path = Path(env["tmp_path"])
    output_path: Path = Path(env["output_path"])

    # tmp putzen, evtl context manager nutzen?
    clear_path(tmp_path)

    # TODO: dynamisch einlesen
    abrechnungsmonat = "2025-08"  # YYYY-MM

    # Aufwandsdaten für den Abrechnungsmonat aus der DB (Excel) laden
    data_path: Path = project_root / env["data_path"]
    db_name: str = CONFIG["db_name"]
    sheet_name: Optional[str] = CONFIG.get("sheet_name")

    invoice_data: pd.DataFrame = load_data(
        data_path / db_name, sheet_name, abrechnungsmonat
    )

    check_data_consistency(invoice_data)

    # Rechnungsübersicht initialisieren
    summary_rows = []

    # Nach Zahlungsdienstleister gruppieren
    zd_grouped = invoice_data.groupby("ZDNR")

    for zdnr, zd_data in zd_grouped:
        ic("Verarbeite Zahlungsdienstleister " + zd_data["ZD_Name"].iloc[0])
        # Liste für die Rechnungsdateien dieses Zahlungsdienstleisters
        invoice_group: list[Path] = []

        # Innerhalb des ZD nach Klient-Nr. gruppieren
        client_grouped = zd_data.groupby("Klient-Nr.")

        # Für jeden Klienten eine Rechnung erstellen
        for client_id, client_details in client_grouped:
            formatted_invoice, updated_data = format_invoice(
                client_details,
            )
            re_nr: str = updated_data["Rechnungsnummer"].iloc[0]
            docx_path: Path = tmp_path / f"Rechnung_{re_nr}.docx"
            formatted_invoice.save(docx_path)

            docx_to_pdf(docx_path, _ := docx_path.with_suffix(".pdf"))

            invoice_group.append(_)

            # Summarische Daten sammeln
            summary_rows.append(
                {
                    "Rechnungsnummer": re_nr,
                    "Klient-Nr.": client_id,
                    "ZDNR": zdnr,
                    "ZD_Name": updated_data["ZD_Name"].iloc[0],
                    "Summe_Kosten": updated_data["Summe_Kosten"].iloc[0],
                    "Rechnungsdatum": updated_data["Rechnungsdatum"].iloc[0],
                }
            )

        merge_pdfs(invoice_group, str(zdnr), abrechnungsmonat)
        # Ende Klienten-Schleife

        # TODO: Wieder rausnehmen, wenn mehrere ZD unterstützt werden
        break

    # Summarische Tabelle als Excel ausgeben
    create_summary(summary_rows, abrechnungsmonat)

    # Alle Rechnungen im Word Format in eine ZIP-Datei packen und im output
    # Verzeichnis ablegen
    zip_docs(tmp_path, output_path / f"Rechnungen_{abrechnungsmonat}.zip")

    # Ende Zahlungsdienstleister-Schleife


if __name__ == "__main__":
    ic.configureOutput(prefix="", outputFunction=print)
    # ic.disable()  # Falls du die Ausgabe komplett abschalten willst

    # Nur das Ergebnis ausgeben (ohne Variable, Datei, Zeilennummer etc.)
    def ic_simple_format(*args):
        return " ".join(str(arg) for arg in args)

    ic.format = ic_simple_format

    main()
