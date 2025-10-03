from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError

class ReportingFactoryConfig(BaseModel):
    """
    Pydantic-Modell für die Konfiguration der ReportingFactory.
    Sorgt für Typsicherheit und Validierung der Konfigurationsdaten.
    Das Passwort für den Blattschutz wird nicht mehr im Code gespeichert,
    sondern sicher aus der Umgebung geladen.
    """

    reporting_template: str = "zeiterfassunsboegen.xlsx"  # Standard-Template-Dateiname
    sheet_password: Optional[str] = None  # Wird im Konstruktor aus der Umgebung geladen


class ReportingFactory:
    """
    Factory-Klasse zur Erstellung von Reporting-Sheets.
    Erwartet ein Pydantic-Modell für die Konfiguration.
    """

    def __init__(self, config: ReportingFactoryConfig):
        """
        Konstruktor erwartet ein Pydantic-Modell für die Konfiguration.
        Das sorgt für Typsicherheit und Validierung der Konfigurationsdaten.
        Das Passwort für den Blattschutz wird sicher aus der Umgebung geladen.
        """
        from module.config import Config

        self.config: ReportingFactoryConfig = config
        # sheet_password sicher aus Umgebungsvariable/.env laden, falls nicht gesetzt
        if not self.config.sheet_password:
            # Versuche zuerst entschlüsseltes Secret zu laden
            secret = Config().get_decrypted_secret("SHEET_PASSWORD_ENC")
            if not secret:
                # Fallback: Unverschlüsseltes Secret
                secret = Config().get_secret("SHEET_PASSWORD")
            if not secret:
                raise RuntimeError(
                    "SHEET_PASSWORD_ENC oder SHEET_PASSWORD nicht gesetzt! Bitte .env anlegen und Passwort eintragen."
                )
            self.config.sheet_password = secret

    def create_reporting_sheet(
        self,
        row: pd.Series,
        reporting_month_dt: datetime,
        output_path: Path,
        template_path: Path,
    ) -> str:
        """
        Erstellt ein Reporting-Sheet auf Basis der übergebenen Datenreihe und speichert es ab.
        Alle Konfigurationswerte werden typisiert über das Pydantic-Modell bezogen.

        Args:
            row (pd.Series): Datenzeile mit den auszufüllenden Werten.
            reporting_month_dt (datetime): Berichtsmonat als Datum.
            output_path (Path): Zielverzeichnis für das Reporting-Sheet.
            template_path (Path): Verzeichnis mit dem Excel-Template.

        Returns:
            str: Dateiname der erzeugten Excel-Datei.
        """
        # Zugriff auf Template-Name und Passwort über das Pydantic-Modell
        template_name: str = self.config.reporting_template

        wb = load_workbook(template_path / template_name)
        ws = wb.active
        if ws is None:
            raise RuntimeError("Kein aktives Arbeitsblatt im Template gefunden.")

        # Blattschutz deaktivieren, um Felder zu beschreiben
        if hasattr(ws, "protection") and ws.protection:
            ws.protection.sheet = False

        # Ausfüllen der relevanten Felder im Excel-Sheet
        try:
            ws["c5"] = row["Sozialpädagogin"]
            ws["g5"] = row["MA_ID"]
            ws["c6"] = reporting_month_dt
            ws["c6"].number_format = "MM.YYYY"
            ws["c7"] = row["Stunden pro Monat"]
            ws["g7"] = row["SPF / BBT"]
            ws["c8"] = row["Kürzel"]
            ws["g8"] = row["KlientNr"]
        except Exception as e:
            raise RuntimeError(f"Fehler beim Ausfüllen des Sheets: {e}")

        # Blattschutz wieder aktivieren und mit Passwort versehen
        if hasattr(ws, "protection") and ws.protection:
            ws.protection.sheet = True
            ws.protection.enable()
            if self.config.sheet_password is None:
                raise RuntimeError("Sheet-Passwort ist nicht gesetzt!")
            ws.protection.set_password(str(self.config.sheet_password))
            # Die folgenden Attribute sind optional und nicht in allen openpyxl-Versionen verfügbar
            for attr, value in [
                ("enable_select_locked_cells", False),
                ("enable_select_unlocked_cells", True),
                ("format_cells", False),
                ("format_columns", False),
                ("format_rows", False),
                ("insert_columns", False),
                ("insert_rows", False),
                ("insert_hyperlinks", False),
                ("delete_columns", False),
                ("delete_rows", False),
                ("sort", False),
                ("auto_filter", False),
                ("objects", False),
                ("scenarios", False),
            ]:
                if hasattr(ws.protection, attr):
                    setattr(ws.protection, attr, value)

        # Dateinamen generieren und Datei speichern
        dateiname: str = f"Aufwandserfassung_{reporting_month_dt.strftime('%Y-%m')}_{row['Kürzel']}.xlsx"
        wb.save(output_path / dateiname)
        return dateiname


# Beispiel für die Initialisierung mit Pydantic
if __name__ == "__main__":
    import yaml

    # Beispiel: YAML-Konfiguration laden und mit Pydantic validieren
    config_path = Path("wegpiraten_reporting_factory_config.yaml")
    with open(config_path, "r") as f:
        raw_config = yaml.safe_load(f)
    try:
        config = ReportingFactoryConfig(**raw_config)
    except ValidationError as e:
        print(f"Konfigurationsfehler: {e}")
        exit(1)

    # Beispielhafte Nutzung
    factory = ReportingFactory(config)
    # Hier müssten row, reporting_month_dt, output_path, template_path übergeben werden
    # factory.create_reporting_sheet(row, reporting_month_dt, output_path, template_path)
