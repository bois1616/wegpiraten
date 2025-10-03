from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from pydantic import ValidationError
from .reporting_config import ReportingConfig  # ReportingConfig importieren


# Pydantic-Modell für die Reporting-Konfiguration


class ReportingProcessor:
    def get_db_password(self) -> str:
        """
        Beispiel: Holt das Datenbank-Passwort sicher aus der Umgebung (z. B. für SQL- oder API-Logins).
        Niemals im Code speichern!
        Rückgabe: Passwort als String oder Exception, falls nicht gesetzt.
        """
        from module.config import Config
        config = Config()
        pw = config.get_secret("DB_PASSWORD")
        if not pw:
            raise RuntimeError("Datenbank-Passwort nicht gesetzt! Bitte .env anlegen und DB_PASSWORD eintragen.")
        return pw

    """
    Klasse zur Verarbeitung von Reporting-Daten.
    Erwartet ausschließlich Pydantic-Modelle für Konfiguration und nutzt Typsicherheit.
    """
    def __init__(self, config: ReportingConfig, factory: object):
        """
        Konstruktor erwartet ein Pydantic-Modell für die Konfiguration.
        Das sorgt für Typsicherheit und Validierung der Konfigurationsdaten.
        Args:
            config (ReportingConfig): Validierte Reporting-Konfiguration.
            factory (object): Factory-Objekt zur Erstellung der Reporting-Sheets.
        """
        self.config: ReportingConfig = config
        self.factory = factory

    def load_client_data(self, reporting_month: str) -> pd.DataFrame:
        """
        Lädt die Klientendaten für den angegebenen Berichtsmonat.
        Nutzt die validierte Pydantic-Konfiguration für alle Pfadangaben.

        Args:
            reporting_month (str): Monat im Format "YYYY-MM".

        Returns:
            pd.DataFrame: Gefilterte Klientendaten.
        """
        # Zugriff auf die Konfigurationsdaten über das Pydantic-Modell
        prj_root = Path(self.config.structure.prj_root)
        data_path = prj_root / self.config.structure.data_path
        db_name = self.config.db_name
        source = data_path / db_name
        table_name = self.config.client_sheet_name

        reporting_month_dt = datetime.strptime(reporting_month, "%Y-%m")
        wb = load_workbook(source, data_only=True)
        for ws in wb.worksheets:
            if table_name in ws.tables:
                table = ws.tables[table_name]
                ref = table.ref
                data = ws[ref]
                rows = [[cell.value for cell in row] for row in data]
                df = pd.DataFrame(rows[1:], columns=rows[0])
                break
        else:
            raise ValueError(f"Tabelle {table_name} nicht gefunden in {db_name}")

        # Datumsfilterung: Nur Klienten, deren "Ende" leer ist oder nach dem Berichtsmonat liegt
        df["Ende"] = pd.to_datetime(df["Ende"], format="%d.%m.%Y", errors="coerce")
        df = df[(df["Ende"].isna()) | (df["Ende"] >= reporting_month_dt)]
        return df

    def run(self, reporting_month: str, output_path: Path, template_path: Path) -> None:
        """
        Führt die Berichtsverarbeitung für den angegebenen Monat aus.

        Args:
            reporting_month (str): Monat im Format "YYYY-MM".
            output_path (Path): Zielverzeichnis für die erzeugten Dateien.
            template_path (Path): Verzeichnis mit den Excel-Templates.
        """
        reporting_month_dt = datetime.strptime(reporting_month, "%Y-%m")
        df = self.load_client_data(reporting_month)
        for idx, row in df.iterrows():
            dateiname = self.factory.create_reporting_sheet(
                row, reporting_month_dt, output_path, template_path
            )
            print(
                f"Erstelle AZ Erfassungsbogen für {row['Sozialpädagogin']} "
                f"({row['Kürzel']}, Ende: {row['Ende']}) -> {dateiname}"
            )

# Beispiel für die Initialisierung mit Pydantic
if __name__ == "__main__":
    import yaml

    # Beispiel: YAML-Konfiguration laden und mit Pydantic validieren
    config_path = Path("wegpiraten_reporting_config.yaml")
    with open(config_path, "r") as f:
        raw_config = yaml.safe_load(f)
    try:
        config = ReportingConfig(**raw_config)
    except ValidationError as e:
        print(f"Konfigurationsfehler: {e}")
        exit(1)

    # Factory-Objekt muss bereitgestellt werden
    factory = None  # Platzhalter
    processor = ReportingProcessor(config, factory)
    # processor.run("2025-08", Path("output"), Path("template.xlsx"))