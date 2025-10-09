from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List, Optional

import pandas as pd
import sqlite3
from loguru import logger
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import ValidationError

from shared_modules.config import Config
from shared_modules.utils import ensure_dir, derive_table_range
from pydantic_models.config.config_data import TimeSheetHeaderCells, TimeSheetRowMapping
from pydantic_models.data.header_data_model import HeaderDataModel
from pydantic_models.config.entity_model_config import EntityModelConfig


class TimeSheetFactory:
    """
    Factory zum Erzeugen einzelner Arbeitszeiterfassungs-Sheets.
    Nutzt das zentrale Config-Objekt und das statische HeaderDataModel.
    """

    def __init__(self, config: Config) -> None:
        """
        Initialisiert die Factory:
        - Prüft einmal die relevanten Config-Werte (Pfad, Template, DB).
        - Baut das Header-Pydantic-Modell dynamisch anhand der Entity-Definition.
        - Lädt (optional entschlüsselt) das Sheet-Passwort.
        """
        self.config: Config = config

        self._validate_header_model()

        self.prj_root: Path = Path(self.config.structure.prj_root)
        self.data_dir: Path = ensure_dir(
            self.prj_root / (getattr(self.config.structure, "local_data_path", None) or "data")
        )
        self.output_dir: Path = ensure_dir(
            self.prj_root / (getattr(self.config.structure, "output_path", None) or "output")
        )
        self.template_dir: Path = ensure_dir(
            self.prj_root / (getattr(self.config.structure, "template_path", None) or "templates")
        )

        self.db_path: Path = self.data_dir / self.config.database.sqlite_db_name
        self.template_file: Path = self.template_dir / self.config.templates.reporting_template

        self.sheet_password: Optional[str] = getattr(self.config, "sheet_password", None)
        if not self.sheet_password:
            secret = self.config.get_decrypted_secret("SHEET_PASSWORD_ENC") or self.config.get_secret("SHEET_PASSWORD")
            if not secret:
                logger.error("SHEET_PASSWORD_ENC oder SHEET_PASSWORD muss gesetzt sein.")
                raise RuntimeError("Excel-Blattschutz-Passwort fehlt.")
            self.sheet_password = secret

        templates_cfg = self.config.templates
        self.sheet_name: Optional[str] = getattr(templates_cfg, "time_sheet_sheet_name", None)
        self.header_cells: TimeSheetHeaderCells = templates_cfg.time_sheet_header_cells
        self.row_mapping: TimeSheetRowMapping = templates_cfg.time_sheet_row_mapping
        self.table_first_col, self.table_start_row, self.table_last_col, self.table_end_row = derive_table_range(
            templates_cfg.time_sheet_data_start_cell,
            templates_cfg.time_sheet_data_end_cell,
        )

    def _validate_header_model(self) -> None:
        """
        Prüft einmalig, ob HeaderDataModel mit der Entity-Definition 'client' kompatibel ist.
        """
        entity_cfg: EntityModelConfig | None = self.config.models.get("client")
        if entity_cfg is None:
            logger.error("Entity 'client' fehlt in der Config.")
            raise ValueError("Entity 'client' fehlt in der Config.")

        entity_fields = {field.name for field in entity_cfg.fields}
        required_fields = {"client_id", "employee_id", "service_type", "short_code", "allowed_hours_per_month"}

        if not required_fields.issubset(entity_fields):
            missing = required_fields - entity_fields
            logger.error(f"Pflichtfelder fehlen in Entity 'client': {missing}")
            raise ValueError(f"Pflichtfelder fehlen in Entity 'client': {missing}")

        model_fields = set(HeaderDataModel.model_fields)

        extra_in_model = model_fields - entity_fields - {
            "client_first_name",
            "client_last_name",
            "employee_first_name",
            "employee_last_name",
        }
        if extra_in_model:
            logger.error(f"HeaderDataModel enthält unbekannte Felder: {extra_in_model}")
            raise ValueError(f"HeaderDataModel enthält unbekannte Felder: {extra_in_model}")

    # --------------------------------------------------------------------- #
    # Datenbeschaffung
    # --------------------------------------------------------------------- #

    def get_db_connection(self) -> sqlite3.Connection:
        """
        Öffnet eine SQLite-Verbindung (Kontextmanager-kompatibel).
        """
        logger.debug(f"Öffne SQLite-DB: {self.db_path}")
        return sqlite3.connect(self.db_path)

    def fetch_reporting_data(self, reporting_month: str) -> List[HeaderDataModel]:
        """
        Lädt alle im Erfassungsmonat aktiven Clients samt Mitarbeiterdaten
        und validiert sie gegen das dynamische Header-Modell.
        """
        month_start = f"{reporting_month}-01"
        sql = """
        SELECT
            c.client_id,
            c.short_code,
            c.allowed_hours_per_month,
            c.employee_id,
            c.first_name AS client_first_name,
            c.last_name AS client_last_name,
            c.service_type,
            e.first_name AS employee_first_name,
            e.last_name AS employee_last_name
        FROM clients c
        LEFT JOIN employees e ON c.employee_id = e.emp_id
        WHERE (c.end_date IS NULL OR c.end_date >= ?)
        """

        logger.info(f"Lese Client-Daten für Monat {reporting_month}.")
        with self.get_db_connection() as conn:
            df = pd.read_sql_query(sql, conn, params=(month_start,))
        logger.info(f"{len(df)} relevante Datensätze für Zeiterfassungs-Sheets geladen.")

        headers: List[HeaderDataModel] = []
        for idx, row in df.iterrows():
            try:
                headers.append(HeaderDataModel(**row.to_dict()))
            except ValidationError as exc:
                logger.error(f"Ungültige Reporting-Daten in Zeile {idx}: {exc}")
        return headers

    # --------------------------------------------------------------------- #
    # Sheet-Erstellung
    # --------------------------------------------------------------------- #

    def create_reporting_sheet(
        self,
        header_data: HeaderDataModel,
        reporting_month_dt: datetime,
        output_path: Optional[Path] = None,
        template_path: Optional[Path] = None,
        sheet_password: Optional[str] = None,
    ) -> Path:
        """
        Erstellt ein Reporting-Sheet auf Basis der übergebenen Kopf-Daten
        und speichert es im Zielverzeichnis.

        Args:
            header_data: Validierte Kopf-Daten (Pydantic-Modell).
            reporting_month_dt: Monatsdatum (erster Tag genügt).
            output_path: Zielordner für das erzeugte Sheet (Default: Config.output_path).
            template_path: Verzeichnis mit dem Template (Default: Config.template_path).
            sheet_password: Optionales Passwort; sonst wird das hinterlegte verwendet.

        Returns:
            Path: Vollständiger Pfad zur erzeugten Datei.
        """
        target_output = ensure_dir(output_path or self.output_dir)
        template_dir = template_path or self.template_dir
        template_file = template_dir / self.config.templates.reporting_template
        assert template_file.exists(), f"Template-Datei nicht gefunden: {template_file}"

        try:
            wb: Workbook = load_workbook(template_file)
            if self.sheet_name:
                if self.sheet_name not in wb.sheetnames:
                    raise RuntimeError(f"Sheet '{self.sheet_name}' fehlt im Template.")
                ws = wb[self.sheet_name]
            else:
                ws = wb.active
        except Exception as exc:
            logger.error(f"Fehler beim Laden des Templates {template_file.name}: {exc}")
            raise RuntimeError(f"Fehler beim Laden des Templates: {exc}") from exc

        cells = self.header_cells
        ws[cells.employee_name] = f"{header_data.employee_first_name or ''} {header_data.employee_last_name or ''}".strip()
        ws[cells.emp_id] = header_data.employee_id
        ws[cells.reporting_month] = reporting_month_dt
        ws[cells.reporting_month].number_format = "MM.YYYY"
        ws[cells.allowed_hours_per_month] = header_data.allowed_hours_per_month
        ws[cells.service_type] = header_data.service_type
        ws[cells.short_code] = header_data.short_code
        ws[cells.client_id] = header_data.client_id

        original_sheet_protected = bool(getattr(ws.protection, "sheet", False))
        if original_sheet_protected:
            ws.protection.sheet = False

        if original_sheet_protected:
            password = sheet_password or self.sheet_password
            if not password:
                raise RuntimeError("Sheet-Passwort ist nicht gesetzt!")
            ws.protection.sheet = True
            ws.protection.set_password(str(password))
            ws.protection.enable()
            # Restriktive Schutz-Einstellungen
            for attr, value in [
                ("enable_select_locked_cells", False),
                ("enable_select_unlocked_cells", True),
                ("format_cells", False),
                ("format_columns", False),
                ("format_rows", False),
                ("insert_columns", False),
                ("insert_rows", False),
                ("insert_hyperlinks", False),
                ("delete_columns", False),
                ("delete_rows", False),
                ("sort", False),
                ("auto_filter", False),
                ("objects", False),
                ("scenarios", False),
            ]:
                if hasattr(ws.protection, attr):
                    setattr(ws.protection, attr, value)

        filename = f"{header_data.client_id} ({header_data.short_code})_{reporting_month_dt.strftime('%Y-%m')}.xlsx"
        target_file = target_output / filename
        try:
            wb.save(target_file)
        except Exception as exc:
            logger.error(f"Fehler beim Speichern der Datei {target_file.name}: {exc}")
            raise RuntimeError(f"Fehler beim Speichern der Datei: {exc}") from exc

        logger.info(f"Reporting-Sheet gespeichert: {target_file}")
        return target_file


if __name__ == "__main__":
    config_path = Path(__file__).parents[3] / ".config" / "wegpiraten_config.yaml"
    config = Config(config_path)
    factory = TimeSheetFactory(config)
    reporting_month = "2025-10"
    reporting_rows = factory.fetch_reporting_data(reporting_month)
    if reporting_rows:
        new_sheet = factory.create_reporting_sheet(
            header_data=reporting_rows[0],
            reporting_month_dt=datetime.strptime(reporting_month, "%Y-%m"),
        )
        logger.info(f"Reporting-Sheet {new_sheet.name} erfolgreich erstellt.")
