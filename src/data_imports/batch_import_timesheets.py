# importiere ein Batch von reporting sheets in die Datenbank
# nutzt die config.py für die Pfade
# Die Daten werden in service_data gespeichert (Rohdaten für die Rechnungserstellung)
# Nutzt reporting_row_model.py für die Datenstruktur
# #
# Gemeinsames Profil: Die Klassen HeaderCells, RowMapping, TableRange und
# ReportingImportProfile kapseln Zellen- und Spaltenbezüge sowie den Datenbereich.
# Diese Strukturen sollten idealerweise in ein gemeinsames Modul
# (z.B. zeiterfassungen/modules/reporting_sheet_profile.py) ausgelagert und
# sowohl in ReportingFactory (Erstellung) als auch hier (Import) genutzt werden.
# So bleibt die Logik synchron.
# Konsistenz mit create_reporting_sheet:
# Die verwendeten Header-Zellen (C5, G5, C6, C7, G7, C8, G8) entsprechen den in
# create_reporting_sheet beschriebenen Feldern.
# Bei Änderungen am Template muss nur HeaderCells angepasst werden.
# Robustheit:
# Einmalige Prüfungen (DB-Pfad, Sheet-Existenz, notwendige Header-Felder)
# mit assert (“prüfe einmal und dann traue”).
# Zeilen werden defensiv konvertiert und validiert.
# Transaktionale Inserts mit Rollback bei Fehler.
# Pfade: Vorerst wird ein fixer Windows-Pfad als Quelle verwendet.
# Wenn dieser nicht existiert, wird auf einen dynamischen Pfad aus der Config
# (structure.imports_path) oder schließlich prj_root/data_imports zurückgefallen.
# Die verarbeiteten Dateien werden in ein Unterverzeichnis importiert verschoben,
#  welches bei Bedarf angelegt wird.
# Pydantic:
# Konsequent für Profil und Datenmodelle. Erweiterbar um weitere Felder (z.B. Notizen aus Spalte G/H).
# Weiter normalisieren?
# Ja, sinnvoll wäre eine Normalisierung der Positionen in eine eigene Tabelle
# timesheet_entries, wenn service_data bereits Rohdaten abbildet.
#  Alternativ kann service_data als staging genutzt werden und ein nachgelagerter
# Prozess normalisiert in ein Faktentableau.

from __future__ import annotations

import re
import sqlite3
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, cast

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ValidationError

from pydantic_models.data.invoice_row_model import InvoiceRowModel
from pydantic_models.data.row_mapping import RowMapping
from pydantic_models.data.timesheet_import_profile import TimeSheetImportProfile
from shared_modules.config import Config
from shared_modules.month_period import MonthPeriod, get_month_period
from shared_modules.utils import (
    choose_existing_path,
    ensure_dir,
    to_date,
    to_float,
    to_year_month_str,
)


class ImportedRowExport(BaseModel):
    """
    Erweiterung des InvoiceRowModel um Kontextinformationen für den Sammel-Export.
    """

    reporting_month: str
    source_file: str
    tenant_id: Optional[str] = None
    client_id: str
    employee_id: str
    service_date: date
    service_type: str
    travel_time: float
    travel_distance: float = 0.0
    direct_time: float
    indirect_time: float
    billable_hours: float
    notes: Optional[str] = None
    hourly_rate: Optional[float] = None
    total_hours: Optional[float] = None
    total_costs: Optional[float] = None


class ImportErrorEntry(BaseModel):
    """Ein einzelner Fehlerfall für das Import-Fehlerprotokoll."""

    timestamp: str
    source_file: str
    category: str
    message: str
    row_number: Optional[int] = None


class TimeSheetsImporter:
    """
    Liest ausgefüllte Aufwandserfassungs-Sheets, validiert jede Positionszeile
    gegen InvoiceRowModel, schreibt sie in service_data und exportiert zusätzlich
    eine Sammeldatei der importierten Daten.
    """

    _INSERT_FIELDS: tuple[str, ...] = (
        "client_id",
        "tenant_id",
        "employee_id",
        "service_date",
        "service_type",
        "travel_time",
        "travel_distance",
        "direct_time",
        "indirect_time",
        "notes",
        "source_file",
        "reporting_month",
    )

    _HEADER_LABELS_WITH_KM: Dict[str, str] = {
        "service_time_col": "uhrzeit",
        "service_date_col": "datum",
        "travel_time_col": "fahrtzeit",
        "travel_distance_col": "km",
        "direct_time_col": "direkter fallkontakt",
        "indirect_time_col": "indirekte fallbearbeitung",
        "billable_hours_col": "stunden",
        "notes_col": "notizen",
    }

    _HEADER_LABELS_NO_KM: Dict[str, str] = {
        "service_time_col": "uhrzeit",
        "service_date_col": "datum",
        "travel_time_col": "fahrtzeit",
        "travel_distance_col": "direkter fallkontakt",
        "direct_time_col": "indirekte fallbearbeitung",
        "indirect_time_col": "stunden",
        "billable_hours_col": "notizen",
    }
    _SHORT_DATE_PATTERN = re.compile(r"^\s*(\d{1,2})\s*[.\-/]\s*(\d{1,2})\s*[.\-/]?\s*$")

    def __init__(self, config: Config, profile: Optional[TimeSheetImportProfile] = None):
        self.config = config
        self.profile = profile or self._build_profile_from_config(config)
        self.masterdata_stem = Path(self.config.database.db_name or "").stem

        prj_root = Path(self.config.structure.prj_root)
        data_dir = prj_root / (getattr(self.config.structure, "local_data_path", None) or "data")
        self.db_path = data_dir / self.config.database.sqlite_db_name
        self.output_dir = ensure_dir(prj_root / (getattr(self.config.structure, "output_path", None) or "output"))
        self.log_dir = ensure_dir(prj_root / (getattr(self.config.structure, "log_path", None) or ".logs"))

        cfg_imports_path = getattr(self.config.structure, "imports_path", None) or self.config.get(
            "structure.imports_path", None
        )
        cfg_done_path = getattr(self.config.structure, "done_path", None) or self.config.get(
            "structure.done_path", None
        )
        default_import = prj_root / "import"
        fallback_local = prj_root / "data_imports"
        default_windows = Path(r"C:\Users\micro\OneDrive\Shared\Beatus\Wegpiraten Unterlagen\data_imports")

        # Kandidatenliste: nimm den ersten existierenden Pfad
        candidates: List[Optional[Path]] = [
            Path(cfg_imports_path) if cfg_imports_path else None,
            default_import,
            fallback_local,
            default_windows,
        ]
        self.source_dir = ensure_dir(choose_existing_path(candidates, default_import))
        self.done_dir = ensure_dir(Path(cfg_done_path) if cfg_done_path else (prj_root / "done"))

        logger.info(f"DB: {self.db_path}")
        logger.info(f"Quelle: {self.source_dir}")
        logger.info(f"Done-Ordner: {self.done_dir}")
        logger.info(f"Fehler-Log-Ordner: {self.log_dir}")
        logger.info("Zeiterfassung-Import: Zeitwerte werden als Minuten gelesen und in Stunden gespeichert.")
        logger.info(
            f"Sheet: {self.profile.sheet_name}, Range: "
            f"{self.profile.table_range.start_row}-{self.profile.table_range.end_row}"
        )

        self._error_entries: List[ImportErrorEntry] = []
        self._valid_client_ids: set[str] = set()
        self._valid_employee_ids: set[str] = set()

        self._ensure_service_data_table()
        self._refresh_fk_cache()

    @staticmethod
    def _build_profile_from_config(config: Config) -> TimeSheetImportProfile:
        return TimeSheetImportProfile.from_config(config.templates)

    def _record_error(
        self,
        source_file: str,
        category: str,
        message: str,
        row_number: Optional[int] = None,
    ) -> None:
        entry = ImportErrorEntry(
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            source_file=source_file,
            category=category,
            message=message,
            row_number=row_number,
        )
        self._error_entries.append(entry)
        row_suffix = f" (Zeile {row_number})" if row_number is not None else ""
        logger.error("{}{}: {}", source_file, row_suffix, message)

    def _record_warning(
        self,
        source_file: str,
        category: str,
        message: str,
        row_number: Optional[int] = None,
    ) -> None:
        entry = ImportErrorEntry(
            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            source_file=source_file,
            category=f"Warnung/{category}",
            message=message,
            row_number=row_number,
        )
        self._error_entries.append(entry)
        row_suffix = f" (Zeile {row_number})" if row_number is not None else ""
        logger.warning("{}{}: {}", source_file, row_suffix, message)

    def _fetch_reference_values(self, table: str, column: str) -> set[str]:
        sql = f"SELECT {column} FROM {table} WHERE {column} IS NOT NULL"
        values: set[str] = set()
        try:
            with sqlite3.connect(self.db_path) as conn:
                for row in conn.execute(sql):
                    value = row[0]
                    if value is None:
                        continue
                    normalized = str(value).strip()
                    if normalized:
                        values.add(normalized)
        except sqlite3.OperationalError as exc:
            self._record_error(
                "SYSTEM",
                "Referenzdaten",
                f"Referenztabelle {table}.{column} nicht lesbar: {exc}",
            )
        return values

    def _refresh_fk_cache(self) -> None:
        self._valid_client_ids = self._fetch_reference_values("clients", "client_id")
        self._valid_employee_ids = self._fetch_reference_values("employees", "emp_id")

    def _validate_header_foreign_keys(self, header: Dict[str, object], source_file: str) -> bool:
        client_id = str(header.get("client_id") or "").strip()
        if not client_id:
            self._record_error(source_file, "Header", "client_id fehlt im Header.")
            return False
        if client_id not in self._valid_client_ids:
            self._record_error(
                source_file,
                "FK-Fehler",
                f"client_id '{client_id}' existiert nicht in den Stammdaten (clients.client_id).",
            )
            return False

        employee_id = str(header.get("employee_id") or "").strip()
        if employee_id and employee_id not in self._valid_employee_ids:
            self._record_error(
                source_file,
                "FK-Fehler",
                f"employee_id '{employee_id}' existiert nicht in den Stammdaten (employees.emp_id).",
            )
            return False

        service_type = str(header.get("service_type") or "").strip()
        if not service_type:
            self._record_error(
                source_file,
                "FK-Fehler",
                "service_type konnte für den Client nicht aus den Stammdaten ermittelt werden.",
            )
            return False
        return True

    def _write_error_report(self) -> Optional[Path]:
        if not self._error_entries:
            return None
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        report_path = self.log_dir / f"timesheet_import_fehler_{stamp}.md"
        report_xlsx_path = self.log_dir / f"timesheet_import_fehler_{stamp}.xlsx"
        lines: List[str] = [
            "# Fehlerprotokoll Timesheet-Import",
            "",
            f"- Zeitpunkt: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"- Anzahl Fehler: {len(self._error_entries)}",
            "",
            "| Zeitpunkt | Datei | Zeile | Kategorie | Fehler |",
            "|---|---|---:|---|---|",
        ]
        for entry in self._error_entries:
            row_number = str(entry.row_number) if entry.row_number is not None else "-"
            safe_message = entry.message.replace("\n", " ").replace("|", "/")
            lines.append(
                f"| {entry.timestamp} | {entry.source_file} | {row_number} | {entry.category} | {safe_message} |"
            )
        report_path.write_text("\n".join(lines), encoding="utf-8")
        df_errors = pd.DataFrame(
            [
                {
                    "zeitpunkt": entry.timestamp,
                    "datei": entry.source_file,
                    "zeile": entry.row_number,
                    "kategorie": entry.category,
                    "fehler": entry.message,
                }
                for entry in self._error_entries
            ]
        )
        with pd.ExcelWriter(report_xlsx_path, engine="openpyxl") as writer:
            df_errors.to_excel(writer, sheet_name="fehlerprotokoll", index=False)

        logger.warning("Fehlerprotokoll geschrieben: {}", report_path)
        logger.warning("Fehlerprotokoll geschrieben: {}", report_xlsx_path)
        return report_path

    def discover_excel_files(self) -> List[Path]:
        files: List[Path] = []
        for path in sorted(p for p in self.source_dir.glob("*.xlsx") if p.is_file()):
            if path.name.startswith("~"):
                logger.info(f"Überspringe temporäre Datei im Import: {path.name}")
                continue
            if self.masterdata_stem and path.stem.startswith(self.masterdata_stem):
                logger.info(f"Überspringe Stammdatendatei im Import: {path.name}")
                continue
            files.append(path)
        logger.info(f"{len(files)} Excel-Dateien entdeckt.")
        return files

    def _ensure_service_data_table(self) -> None:
        sql = """
        CREATE TABLE IF NOT EXISTS service_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id TEXT NOT NULL,
            tenant_id TEXT,
            employee_id TEXT,
            service_date TEXT NOT NULL,
            service_type TEXT NOT NULL,
            travel_time INTEGER,
            travel_distance REAL,
            direct_time INTEGER,
            indirect_time INTEGER,
            notes TEXT,
            source_file TEXT,
            reporting_month TEXT,
            FOREIGN KEY (client_id) REFERENCES clients(client_id),
            FOREIGN KEY (employee_id) REFERENCES employees(emp_id)
        )
        """
        drop_legacy_client_date_index_sql = """
        DROP INDEX IF EXISTS idx_service_data_client_date
        """
        drop_legacy_employee_unique_index_sql = """
        DROP INDEX IF EXISTS idx_service_data_client_date_employee
        """
        client_date_index_sql = """
        CREATE INDEX IF NOT EXISTS idx_service_data_client_date
        ON service_data (client_id, service_date)
        """
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON")
            conn.execute(sql)
            columns = {row[1] for row in conn.execute("PRAGMA table_info(service_data)").fetchall()}
            if "tenant_id" not in columns:
                conn.execute("ALTER TABLE service_data ADD COLUMN tenant_id TEXT")
                logger.info("Tabelle service_data um Spalte tenant_id erweitert.")
            # Legacy-Dedup-Tabelle entfernen (verhindert FK-Konflikt beim DELETE FROM service_data)
            conn.execute("DROP TABLE IF EXISTS service_data_import_dedup")
            conn.execute(drop_legacy_client_date_index_sql)
            conn.execute(drop_legacy_employee_unique_index_sql)
            conn.execute(client_date_index_sql)
            conn.commit()

    def _reset_service_data(self) -> None:
        """
        Leert service_data für einen sauberen Batch-Lauf.
        """
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON")
            conn.execute("DELETE FROM service_data")
            conn.commit()
        logger.warning("service_data wurde vor dem Import zurückgesetzt.")

    @staticmethod
    def _normalize_label(value: object) -> str:
        return " ".join(str(value or "").strip().lower().split())

    def _get_header_label(self, ws: Worksheet, column: str) -> str:
        row_idx = self.profile.table_range.start_row
        value = ws[f"{column}{row_idx}"].value
        return self._normalize_label(value)

    def _detect_notes_column(self, ws: Worksheet) -> Optional[str]:
        """Ermittelt die Notizspalte robust über den Header-Text."""
        row_idx = self.profile.table_range.start_row
        for column in ("H", "G", "F"):
            value = self._normalize_label(ws[f"{column}{row_idx}"].value)
            if value == "notizen":
                return column
        return None

    def _resolve_service_type(self, client_id: Optional[str]) -> Optional[str]:
        if not client_id:
            return None
        sql = """
        SELECT st.code
        FROM clients c
        LEFT JOIN service_types st ON c.service_type = st.service_type_id
        WHERE c.client_id = ?
        """
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute(sql, (client_id,)).fetchone()
        if row and row[0]:
            return str(row[0]).strip()
        return None

    def _resolve_hourly_rate(self, client_id: Optional[str]) -> Optional[float]:
        if not client_id:
            return None
        sql = """
        SELECT st.hourly_rate
        FROM clients c
        LEFT JOIN service_types st ON c.service_type = st.service_type_id
        WHERE c.client_id = ?
        """
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute(sql, (client_id,)).fetchone()
        if row and row[0] is not None:
            return to_float(row[0])
        return None

    def _resolve_employee_id(self, client_id: Optional[str]) -> Optional[str]:
        if not client_id:
            return None
        sql = "SELECT employee_id FROM clients WHERE client_id = ?"
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute(sql, (client_id,)).fetchone()
        if row and row[0]:
            return str(row[0]).strip()
        return None

    def _resolve_tenant_id(self, client_id: Optional[str]) -> Optional[str]:
        if not client_id:
            return None
        sql = "SELECT tenant_id FROM clients WHERE client_id = ?"
        with sqlite3.connect(self.db_path) as conn:
            row = conn.execute(sql, (client_id,)).fetchone()
        if row and row[0]:
            return str(row[0]).strip()
        return None

    def _determine_row_mapping(self, ws: Worksheet) -> Optional[tuple[RowMapping, bool]]:
        mp = self.profile.row_mapping
        detected_notes_col = self._detect_notes_column(ws)
        effective_notes_col = mp.notes_col
        if detected_notes_col and detected_notes_col != mp.notes_col:
            logger.info(
                "Notizspalte automatisch erkannt: {} (statt konfiguriert {}).",
                detected_notes_col,
                mp.notes_col,
            )
            effective_notes_col = detected_notes_col
            mp = RowMapping(
                service_time_col=mp.service_time_col,
                service_date_col=mp.service_date_col,
                travel_time_col=mp.travel_time_col,
                travel_distance_col=mp.travel_distance_col,
                direct_time_col=mp.direct_time_col,
                indirect_time_col=mp.indirect_time_col,
                billable_hours_col=mp.billable_hours_col,
                notes_col=effective_notes_col,
            )

        def read_labels(expected: Dict[str, str]) -> Dict[str, str]:
            labels: Dict[str, str] = {}
            for key in expected:
                column = getattr(mp, key, None)
                if not column:
                    continue
                labels[key] = self._get_header_label(ws, column)
            return labels

        def matches(labels: Dict[str, str], expected: Dict[str, str]) -> bool:
            for key, expected_label in expected.items():
                column = getattr(mp, key, None)
                if not column:
                    continue
                if labels.get(key) != expected_label:
                    return False
            return True

        expected = {
            "service_time_col": "uhrzeit",
            "service_date_col": "datum",
            "travel_time_col": "fahrtzeit",
            "direct_time_col": "direkter fallkontakt",
            "indirect_time_col": "indirekte fallbearbeitung",
        }
        notes_label = self._get_header_label(ws, mp.notes_col) if mp.notes_col else ""
        if mp.notes_col and (detected_notes_col or notes_label == "notizen"):
            expected["notes_col"] = "notizen"
        elif mp.notes_col:
            logger.info("Notizspalte im Header nicht gefunden – Import läuft ohne Header-Prüfung für Notizen.")
        if mp.travel_distance_col:
            expected["travel_distance_col"] = "km"
        if mp.billable_hours_col:
            expected["billable_hours_col"] = "stunden"

        labels = read_labels(expected)
        if matches(labels, expected):
            return mp, bool(mp.travel_distance_col)

        if mp.travel_distance_col and mp.billable_hours_col:
            legacy_labels = read_labels(self._HEADER_LABELS_NO_KM)
            if matches(legacy_labels, self._HEADER_LABELS_NO_KM):
                shifted = RowMapping(
                    service_time_col=mp.service_time_col,
                    service_date_col=mp.service_date_col,
                    travel_time_col=mp.travel_time_col,
                    travel_distance_col=mp.travel_distance_col,
                    direct_time_col=mp.travel_distance_col,
                    indirect_time_col=mp.direct_time_col,
                    billable_hours_col=mp.indirect_time_col,
                    notes_col=mp.billable_hours_col,
                )
                return shifted, False

        logger.error(
            "Header-Struktur nicht erkannt. Gefundene Labels: {}",
            labels,
        )
        return None

    def _read_header(self, ws: Worksheet) -> Dict[str, object]:
        cells = self.profile.header_cells
        employee_id = ws[cells.emp_id].value  # type: ignore[union-attr]
        client_id = ws[cells.client_id].value  # type: ignore[union-attr]
        allowed_hours = ws[cells.allowed_hours_per_month].value  # type: ignore[union-attr]

        client_id_str = str(client_id).strip() if client_id is not None else ""
        if not client_id_str:
            client_id_str = ""

        service_type = self._resolve_service_type(client_id_str or None)
        resolved_employee_id = self._resolve_employee_id(client_id_str or None)
        resolved_tenant_id = self._resolve_tenant_id(client_id_str or None)
        resolved_hourly_rate = self._resolve_hourly_rate(client_id_str or None)
        if not employee_id:
            employee_id = resolved_employee_id
        employee_id_str = str(employee_id).strip() if employee_id is not None else ""

        return {
            "employee_fullname": ws[cells.employee_name].value,  # type: ignore[union-attr]
            "employee_id": employee_id_str or None,
            "reporting_month": ws[cells.reporting_month].value,  # type: ignore[union-attr]
            "allowed_hours_per_month": allowed_hours,
            "service_type": service_type,
            "short_code": ws[cells.short_code].value,  # type: ignore[union-attr]
            "client_id": client_id_str or None,
            "tenant_id": resolved_tenant_id,
            "hourly_rate": resolved_hourly_rate,
        }

    def _read_rows(
        self,
        ws: Worksheet,
        header: Dict[str, object],
        row_mapping: RowMapping,
        has_travel_distance: bool,
        reporting_month: str,
        reporting_period: MonthPeriod,
        source_file: str,
    ) -> tuple[List[Dict[str, Any]], bool]:
        """Liest alle Datenzeilen. Gibt (rows, has_fatal_date_error) zurück."""
        rng = self.profile.table_range
        mp = row_mapping

        rows: List[Dict[str, Any]] = []
        has_fatal_date_error = False
        for row_idx in range(rng.start_row, rng.end_row + 1):
            v_date = ws[f"{mp.service_date_col}{row_idx}"].value
            v_travel = ws[f"{mp.travel_time_col}{row_idx}"].value
            v_distance = (
                ws[f"{mp.travel_distance_col}{row_idx}"].value
                if has_travel_distance and mp.travel_distance_col
                else None
            )
            v_direct = ws[f"{mp.direct_time_col}{row_idx}"].value
            v_indirect = ws[f"{mp.indirect_time_col}{row_idx}"].value
            v_billable = ws[f"{mp.billable_hours_col}{row_idx}"].value if mp.billable_hours_col else None
            v_notes = ws[f"{mp.notes_col}{row_idx}"].value if mp.notes_col else None

            service_date = to_date(v_date)
            if service_date is None:
                service_date = self._parse_partial_service_date(v_date, reporting_period.start.year)
                if service_date is not None:
                    logger.info(
                        "Partielles Datum korrigiert in {} Zeile {}: '{}' -> {}",
                        source_file,
                        row_idx,
                        v_date,
                        service_date.isoformat(),
                    )
            travel = int(round(to_float(v_travel) or 0.0))
            distance = to_float(v_distance) or 0.0
            direct = int(round(to_float(v_direct) or 0.0))
            indirect = int(round(to_float(v_indirect) or 0.0))
            billable = int(round(to_float(v_billable) or 0.0)) if v_billable is not None else None

            if service_date is None:
                if (travel + direct + indirect) == 0:
                    continue
                # Zeiten vorhanden, aber Datum nicht identifizierbar → fataler Fehler
                self._record_error(
                    source_file,
                    "Datenfehler",
                    f"Datum nicht identifizierbar bei vorhandenen Zeiten "
                    f"(F={travel}min, D={direct}min, I={indirect}min) – Timesheet wird nicht importiert.",
                    row_number=row_idx,
                )
                has_fatal_date_error = True
                continue

            if not self._is_within_reporting_period(service_date, reporting_period):
                self._record_warning(
                    source_file,
                    "Leistungsdatum",
                    f"Leistungsdatum ausserhalb Abrechnungsmonat: {service_date.isoformat()} (CLI-Monat: {reporting_month})",
                    row_number=row_idx,
                )

            employee_id = header.get("employee_id")
            employee_id_str = str(employee_id).strip() if employee_id is not None else ""
            payload = {
                "client_id": str(header.get("client_id") or "").strip(),
                "employee_id": employee_id_str or "UNBEKANNT",
                "service_date": service_date,
                "service_type": str(header.get("service_type") or "").strip(),
                "travel_time": travel,
                "direct_time": direct,
                "indirect_time": indirect,
                "billable_hours": billable,
            }

            try:
                validated = InvoiceRowModel(**payload)
                rows.append(
                    {
                        "client_id": validated.client_id,
                        "tenant_id": str(header.get("tenant_id") or "").strip() or None,
                        "employee_id": employee_id_str or None,
                        "service_date": validated.service_date,
                        "service_type": validated.service_type,
                        "travel_time": validated.travel_time,
                        "travel_distance": distance,
                        "direct_time": validated.direct_time,
                        "indirect_time": validated.indirect_time,
                        "billable_hours": validated.billable_hours,
                        "hourly_rate": header.get("hourly_rate"),
                        "notes": str(v_notes).strip() if v_notes is not None else None,
                        "source_file": source_file,
                        "reporting_month": reporting_month,
                        "_source_row": row_idx,
                    }
                )
            except ValidationError as exc:
                short_errors = "; ".join(
                    f"{'.'.join(str(loc) for loc in error.get('loc', []))}: {error.get('msg', 'ungültig')}"
                    for error in exc.errors()
                )
                self._record_error(
                    source_file,
                    "Validierung",
                    f"Validierungsfehler: {short_errors or str(exc)}",
                    row_number=row_idx,
                )

        return rows, has_fatal_date_error

    def _row_params(self, record: Dict[str, Any]) -> tuple[Any, ...]:
        data = record.copy()
        if isinstance(data.get("service_date"), date):
            data["service_date"] = data["service_date"].isoformat()
        return tuple(data.get(column) for column in self._INSERT_FIELDS)

    def _parse_partial_service_date(self, raw_value: object, reporting_year: int) -> Optional[date]:
        """
        Interpretiert tt.mm. bzw. t.m. als Datum mit Jahr aus dem CLI-Leistungsmonat.
        """
        if not isinstance(raw_value, str):
            return None
        match = self._SHORT_DATE_PATTERN.match(raw_value)
        if not match:
            return None

        day = int(match.group(1))
        month = int(match.group(2))
        try:
            return date(reporting_year, month, day)
        except ValueError:
            return None

    @staticmethod
    def _is_within_reporting_period(service_date: date, reporting_period: MonthPeriod) -> bool:
        start_date = reporting_period.start.date()
        end_date = reporting_period.end.date()
        return start_date <= service_date <= end_date

    @staticmethod
    def _normalized_time_value(value: object) -> int:
        return int(round(to_float(value) or 0.0))

    def _import_rows(self, rows: Iterable[Dict[str, Any]], source_file: str) -> tuple[int, List[Dict[str, Any]]]:
        insert_sql = f"""
        INSERT INTO service_data (
            {", ".join(self._INSERT_FIELDS)}
        ) VALUES ({", ".join(["?"] * len(self._INSERT_FIELDS))})
        """
        count = 0
        imported_rows: List[Dict[str, Any]] = []
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("PRAGMA foreign_keys = ON")
            cur = conn.cursor()
            for record in rows:
                try:
                    cur.execute(insert_sql, self._row_params(record))
                    count += 1
                    imported_rows.append(record)
                except sqlite3.IntegrityError as exc:
                    self._record_error(
                        source_file,
                        "FK-Fehler",
                        f"Datenbank-Constraint verletzt: {exc}",
                        row_number=record.get("_source_row"),
                    )
                except sqlite3.DatabaseError as exc:
                    self._record_error(
                        source_file,
                        "DB-Fehler",
                        f"Zeile konnte nicht gespeichert werden: {exc}",
                        row_number=record.get("_source_row"),
                    )
            conn.commit()
        return count, imported_rows

    def _move_to_done(self, file_path: Path) -> Path:
        """
        Verschiebt die verarbeitete Datei in das Verzeichnis 'done'.
        Bei Namenskollision wird die bestehende Datei überschrieben.
        """
        target = self.done_dir / file_path.name
        file_path.replace(target)
        return target

    def _remove_sheet_protection(self, file_path: Path) -> None:
        """
        Entfernt den Blattschutz auf allen Tabellenblättern einer Excel-Datei.
        Zusätzlich werden die relevanten Eingabebereiche explizit entsperrt.
        """
        try:
            wb = load_workbook(file_path)
            for ws in wb.worksheets:
                ws.protection.sheet = False
                for row in ws["C5:F8"]:
                    for cell in row:
                        base_protection = copy(cell.protection) if cell.protection is not None else Protection()
                        base_protection.locked = False
                        cell.protection = base_protection
                for row in ws["A12:H29"]:
                    for cell in row:
                        base_protection = copy(cell.protection) if cell.protection is not None else Protection()
                        base_protection.locked = False
                        cell.protection = base_protection
            wb.save(file_path)
            logger.info("Blattschutz entfernt und Zellbereiche entsperrt: {}", file_path.name)
        except Exception as exc:
            self._record_error(
                file_path.name,
                "Datei",
                f"Blattschutz konnte nicht entfernt werden: {exc}",
            )

    # Spaltenreihenfolge für den Detailexport
    _EXPORT_COLUMNS: tuple[str, ...] = (
        "reporting_month",
        "source_file",
        "service_date",
        "tenant_id",
        "client_id",
        "employee_id",
        "service_type",
        "hourly_rate",
        "travel_time",
        "direct_time",
        "indirect_time",
        "notes",
    )

    def _export_rows_to_excel(self, export_rows: List[ImportedRowExport], reporting_month: str) -> Path:
        """
        Schreibt alle importierten Zeilen als Sammeldatei ins Output-Verzeichnis.
        Sheet 1: Detaildaten (nur Spalten aus _EXPORT_COLUMNS)
        Sheet 2: Rechnungsvorschau – aggregiert je Klient mit Stundensumme und Rechnungsbetrag.
        Stellt sicher, dass der Monatsname gesetzt ist; andernfalls wird der Fehler protokolliert
        und als ValueError weitergereicht.
        """
        if not reporting_month:
            logger.error("Export abgebrochen: reporting_month wurde nicht ermittelt.")
            raise ValueError("reporting_month darf für den Export nicht leer sein.")
        out_file = self.output_dir / f"importierte_daten_{reporting_month}.xlsx"

        # --- Detailsheet ---
        df = pd.DataFrame(
            [
                {
                    **row.model_dump(exclude={"service_date"}),
                    "service_date": row.service_date.isoformat(),
                }
                for row in export_rows
            ]
        )
        # Zeitspalten sind bereits in ganzen Minuten gespeichert – nur in int konvertieren.
        for time_column in ("travel_time", "direct_time", "indirect_time"):
            if time_column in df.columns:
                numeric: pd.Series = cast(pd.Series, pd.to_numeric(df[time_column], errors="coerce"))
                df[time_column] = numeric.fillna(0).round().astype(int)
        # Nur definierte Spalten in der gewünschten Reihenfolge exportieren
        export_cols = [c for c in self._EXPORT_COLUMNS if c in df.columns]
        df_detail = df[export_cols]

        # --- Rechnungsvorschau (Pivot je Klient) ---
        df_pivot = self._build_invoice_preview(df)

        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            df_detail.to_excel(writer, sheet_name="importierte_daten", index=False)
            df_pivot.to_excel(writer, sheet_name="rechnungsvorschau", index=False)

        logger.info(f"Export-Datei geschrieben: {out_file}")
        return out_file

    @staticmethod
    def _build_invoice_preview(df: pd.DataFrame) -> pd.DataFrame:
        """
        Aggregiert die Detaildaten je Klient für eine Rechnungsvorschau.
        Gibt einen DataFrame mit einer Zeile pro Klient zurück.
        """
        group_cols = ["client_id"]
        for optional in ("tenant_id", "service_type"):
            if optional in df.columns:
                group_cols.append(optional)

        agg: Dict[str, Any] = {}
        for col in ("travel_time", "direct_time", "indirect_time"):
            if col in df.columns:
                agg[col] = "sum"
        if "hourly_rate" in df.columns:
            agg["hourly_rate"] = "first"

        if not agg:
            return pd.DataFrame()

        pivot = df.groupby(group_cols, dropna=False).agg(agg).reset_index()

        # Gesamtminuten und Stunden berechnen
        time_cols = [c for c in ("travel_time", "direct_time", "indirect_time") if c in pivot.columns]
        if time_cols:
            pivot["total_minutes"] = pivot[time_cols].sum(axis=1).round().astype(int)
            pivot["total_hours"] = (pivot["total_minutes"] / 60).round(2)

        # Rechnungssumme = Stunden * Stundenansatz
        if "hourly_rate" in pivot.columns and "total_hours" in pivot.columns:
            pivot["rechnungssumme_chf"] = (pivot["total_hours"] * pivot["hourly_rate"]).round(2)

        # Spaltenreihenfolge
        ordered = [c for c in group_cols if c in pivot.columns]
        for col in (
            "travel_time",
            "direct_time",
            "indirect_time",
            "total_minutes",
            "total_hours",
            "hourly_rate",
            "rechnungssumme_chf",
        ):
            if col in pivot.columns:
                ordered.append(col)

        return cast(pd.DataFrame, pivot[[c for c in ordered if c in pivot.columns]])

    def process_file(
        self,
        file_path: Path,
        reporting_month: str,
        reporting_period: MonthPeriod,
    ) -> tuple[int, Path, List[ImportedRowExport], str]:
        logger.info(f"Verarbeite: {file_path.name}")
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as exc:
            self._record_error(file_path.name, "Datei", f"Datei konnte nicht gelesen werden: {exc}")
            return 0, file_path, [], reporting_month
        if self.profile.sheet_name:
            if self.profile.sheet_name not in wb.sheetnames:
                self._record_error(
                    file_path.name,
                    "Struktur",
                    f"Sheet '{self.profile.sheet_name}' fehlt in der Datei.",
                )
                return 0, file_path, [], reporting_month
            ws = wb[self.profile.sheet_name]
        else:
            ws = wb.active
            if ws is None:
                self._record_error(file_path.name, "Struktur", "Kein aktives Sheet gefunden.")
                return 0, file_path, [], reporting_month

        header = self._read_header(ws)

        # C6-Abrechnungsmonat gegen CLI-Monat prüfen
        sheet_month_str = to_year_month_str(header.get("reporting_month"))
        if sheet_month_str and sheet_month_str != reporting_month:
            self._record_warning(
                file_path.name,
                "Abrechnungsmonat",
                f"Abrechnungsmonat im Timesheet (C6: {header.get('reporting_month')!r} → {sheet_month_str}) "
                f"weicht vom CLI-Monat ({reporting_month}) ab. Leistungsdaten werden gegen CLI-Monat geprüft.",
            )

        mapping = self._determine_row_mapping(ws)
        if mapping is None:
            self._record_error(file_path.name, "Struktur", "Header-Struktur ungültig.")
            return 0, file_path, [], reporting_month
        row_mapping, has_travel_distance = mapping

        if not self._validate_header_foreign_keys(header, file_path.name):
            logger.warning("Datei {} wegen Header-/FK-Fehlern übersprungen.", file_path.name)
            return 0, file_path, [], reporting_month

        if not header.get("employee_id"):
            logger.warning("employee_id fehlt im Header – Import erfolgt ohne employee_id ({})", file_path.name)

        logger.info(
            "Header: client_id={}, employee_id={}, service_type={}, month={}",
            header.get("client_id"),
            header.get("employee_id"),
            header.get("service_type"),
            reporting_month,
        )

        rows, has_fatal_date_error = self._read_rows(
            ws,
            header,
            row_mapping,
            has_travel_distance,
            reporting_month,
            reporting_period,
            file_path.name,
        )
        if has_fatal_date_error:
            logger.error(
                "Datei {} enthält nicht identifizierbare Datumsangaben – wird nicht importiert und bleibt im Import.",
                file_path.name,
            )
            return 0, file_path, [], reporting_month
        if not rows:
            logger.warning("Keine importierbaren Zeilen in {} gefunden – Datei bleibt im Import.", file_path.name)
            return 0, file_path, [], reporting_month

        imported_count, imported_rows = self._import_rows(rows, file_path.name)
        if imported_count == 0:
            logger.warning("Keine gültigen Zeilen aus {} gespeichert – Datei bleibt im Import.", file_path.name)
            return 0, file_path, [], reporting_month

        total_minutes = sum(
            (r.get("travel_time") or 0) + (r.get("direct_time") or 0) + (r.get("indirect_time") or 0)
            for r in imported_rows
        )
        if total_minutes == 0:
            self._record_warning(
                file_path.name,
                "Zeitsumme",
                "Zeitsumme aller importierten Zeilen ist 0 Minuten (leere oder fehlende Zeitwerte).",
            )

        moved = self._move_to_done(file_path)
        self._remove_sheet_protection(moved)

        export_rows = [
            ImportedRowExport(**{k: v for k, v in row.items() if k != "_source_row"}) for row in imported_rows
        ]

        logger.info(f"{imported_count} Zeilen importiert aus {file_path.name}. Verschoben nach {moved.name}")
        return imported_count, moved, export_rows, reporting_month

    def run(self, reporting_month: str, reset: bool = True) -> int:
        """
        Führt den Batch-Import aus und schreibt eine Sammel-Excel-Datei ins Output-Verzeichnis:
        output/importierte_daten_{reporting_month}.xlsx

        - reporting_month ist verbindlich und wird für Datumsinterpretation/Prüfung genutzt.
        - Wenn reset=True, wird service_data vor dem Lauf geleert.
        - Gibt die Gesamtanzahl importierter Zeilen zurück.
        """
        reporting_period = get_month_period(reporting_month)
        normalized_reporting_month = reporting_period.start.strftime("%Y-%m")
        if reset:
            self._reset_service_data()

        files = self.discover_excel_files()
        total = 0
        all_export_rows: List[ImportedRowExport] = []

        for file_path in files:
            try:
                count, _, export_rows, _ = self.process_file(
                    file_path=file_path,
                    reporting_month=normalized_reporting_month,
                    reporting_period=reporting_period,
                )
                total += count
                all_export_rows.extend(export_rows)
            except ValueError as err:
                self._record_error(file_path.name, "Struktur", str(err))
            except Exception as exc:
                self._record_error(file_path.name, "Unerwartet", f"Datei konnte nicht verarbeitet werden: {exc}")

        logger.info(f"Batch abgeschlossen. Gesamt importiert: {total}")

        if all_export_rows:
            self._export_rows_to_excel(all_export_rows, normalized_reporting_month)
        self._write_error_report()

        return total


def main() -> None:
    """
    Einstiegspunkt:
    - Lädt zentrale Config (Pydantic-validiert).
    - Initialisiert Importer.
    - Führt den Batch-Import aus und schreibt Sammel-Excel ins Output-Verzeichnis.
    """
    config_path = Path(__file__).parents[2] / ".config" / "wegpiraten_config.yaml"
    config = Config(config_path)  # Singleton, lädt und validiert, setzt Logging

    importer = TimeSheetsImporter(config)
    importer.run("2025-10")  # optional: importer.run("2025-10", reset=False)


if __name__ == "__main__":
    main()
