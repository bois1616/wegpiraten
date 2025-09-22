from datetime import datetime
from pathlib import Path

import pandas as pd
from module.config import Config
from module.document_utils import DocumentUtils
from module.entity import PrivatePerson
from openpyxl import load_workbook


def load_client_data(config: Config, abrechnungsmonat: str) -> pd.DataFrame:
    prj_root = Path(config.data["structure"]["prj_root"])
    data_path = prj_root / config.data["structure"]["data_path"]
    db_name = config.data["db_name"]
    table_name = config.data.get("client_table", "MD_Client")

    abrechnungsmonat_dt = datetime.strptime(abrechnungsmonat, "%Y-%m")
    wb = load_workbook(data_path / db_name, data_only=True)
    for ws in wb.worksheets:
        if table_name in ws.tables:
            table = ws.tables[table_name]
            ref = table.ref
            data = ws[ref]
            rows = [[cell.value for cell in row] for row in data]
            df = pd.DataFrame(rows[1:], columns=rows[0])
            break
    else:
        raise ValueError(f"Tabelle {table_name} nicht gefunden in {db_name}")

    df["Ende"] = pd.to_datetime(df["Ende"], format="%d.%m.%Y", errors="coerce")
    df = df[(df["Ende"].isna()) | (df["Ende"] >= abrechnungsmonat_dt)]
    return df

def create_reporting_sheet(config: Config, row: pd.Series, abrechnungsmonat_dt: datetime, output_path: Path, template_path: Path):
    template_name = config.data.get("zeiterfassung_template", "zeiterfassunsboegen.xlsx")
    wb = load_workbook(template_path / template_name)
    ws = wb.active

    # Blattschutz deaktivieren
    ws.protection.sheet = False

    # Felder befüllen
    ws["c5"] = row["Sozialpädagogin"]
    ws["g5"] = row["MA_ID"]
    ws["c6"] = abrechnungsmonat_dt
    ws["c6"].number_format = "MM.YYYY"
    ws["c7"] = row["Stunden pro Monat"]
    ws["g7"] = row["SPF / BBT"]
    ws["c8"] = row["Kürzel"]
    ws["g8"] = row["KlientNr"]

    ws.protection.sheet = True
    ws.protection.enable()
    ws.protection.set_password(config.data.get("sheet_password", "wegpiraten"))
    ws.protection.enable_select_locked_cells = False
    ws.protection.enable_select_unlocked_cells = True
    ws.protection.format_cells = False
    ws.protection.format_columns = False
    ws.protection.format_rows = False
    ws.protection.insert_columns = False
    ws.protection.insert_rows = False
    ws.protection.insert_hyperlinks = False
    ws.protection.delete_columns = False
    ws.protection.delete_rows = False
    ws.protection.sort = False
    ws.protection.auto_filter = False
    ws.protection.objects = False
    ws.protection.scenarios = False
    
    # Neuen Z Erfassungsbogen erstellen
    dateiname = f"Aufwandserfassung_{abrechnungsmonat}_{row['Kürzel']}.xlsx"
    wb.save(output_path / dateiname)
    return dateiname

def main():
    # Konfiguration laden
    config_path = Path(__file__).parent.parent.parent / ".config" / "wegpiraten_config.yaml"
    config = Config()
    config.load(config_path)

    prj_root = Path(config.data["structure"]["prj_root"])
    data_path = prj_root / config.data["structure"]["data_path"]
    output_path = prj_root / config.data["structure"]["output_path"]
    template_path = prj_root / config.data["structure"]["template_path"]

    reporting_month = config.data.get("abrechnungsmonat", "2025-09")
    abrechnungsmonat_dt = datetime.strptime(reporting_month, "%Y-%m")

    df = load_client_data(config, reporting_month)
    print(df)

    for idx, row in df.iterrows():
        dateiname = create_reporting_sheet(
            config, row, abrechnungsmonat_dt, output_path, template_path
        )
        print(f"Erstelle AZ Erfassungsbogen für {row['Sozialpädagogin']} ({row['Kürzel']}, Ende: {row['Ende']}) -> {dateiname}")

if __name__ == "__main__":
    main()