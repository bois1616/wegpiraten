from datetime import datetime
from pathlib import Path
from typing import Optional, List
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from pydantic import ValidationError
from loguru import logger

from shared_modules.config import Config

from .reporting_row_model import ReportingRowModel

class ReportingFactory:
    """
    Factory-Klasse zur Erstellung von Reporting-Sheets.
    Holt alle Einstellungen und Daten aus dem zentralen Config-Objekt.
    """

    def __init__(self, config: Config):
        """
        Initialisiert die Factory mit dem zentralen Config-Objekt.
        Holt das Sheet-Passwort sicher aus der Umgebung, falls nicht in der Config gesetzt.
        """
        self.config: Config = config
        # Sheet-Passwort aus Config oder Umgebungsvariable holen
        self.sheet_password: Optional[str] = getattr(self.config, "sheet_password", None)
        if not self.sheet_password:
            secret = self.config.get_decrypted_secret("SHEET_PASSWORD_ENC")
            if not secret:
                secret = self.config.get_secret("SHEET_PASSWORD")
            if not secret:
                logger.error("SHEET_PASSWORD_ENC oder SHEET_PASSWORD nicht gesetzt! Bitte .env anlegen und Passwort eintragen.")
                raise RuntimeError("SHEET_PASSWORD_ENC oder SHEET_PASSWORD nicht gesetzt! Bitte .env anlegen und Passwort eintragen.")
            self.sheet_password = secret

    def get_db_connection(self, db_path: Path) -> sqlite3.Connection:
        """
        Öffnet eine SQLite-Verbindung zur angegebenen Datei.
        """
        logger.info(f"Öffne SQLite-DB: {db_path}")
        return sqlite3.connect(db_path)

    def fetch_reporting_data(self, db_path: Path, reporting_month: str) -> List[ReportingRowModel]:
        """
        Holt die für die Zeiterfassungs-Sheets benötigten Felder aus der Tabelle clients,
        ergänzt um die relevanten Daten aus employees per JOIN.
        Berücksichtigt nur Clients, die im Erfassungsmonat noch aktiv sind.
        Gibt eine Liste von validierten ReportingRowModel-Instanzen zurück.
        """
        month_start = f"{reporting_month}-01"
        sql = """
        SELECT
            e.first_name AS employee_first_name,
            e.last_name AS employee_last_name,
            c.employee_id,
            c.allowed_hours_per_month,
            c.service_type,
            c.short_code,
            c.client_id
        FROM clients c
        LEFT JOIN employees e ON c.employee_id = e.emp_id
        WHERE (c.end_date IS NULL OR c.end_date >= ?)
        """
        logger.info(f"Führe Reporting-Query für Monat {reporting_month} aus.")
        with self.get_db_connection(db_path) as conn:
            df = pd.read_sql_query(sql, conn, params=(month_start,))
        logger.info(f"{len(df)} relevante Datensätze für Zeiterfassungs-Sheets geladen.")

        reporting_rows: List[ReportingRowModel] = []
        for idx, row in df.iterrows():
            try:
                reporting_row = ReportingRowModel(**row.to_dict())
                reporting_rows.append(reporting_row)
            except ValidationError as e:
                logger.error(f"Ungültige Reporting-Daten in Zeile {idx}: {e}")
        return reporting_rows

    def create_reporting_sheet(
        self,
        header_data: ReportingRowModel,
        reporting_month_dt: datetime,
        output_path: Path,
        template_path: Path,
        sheet_password: Optional[str] = None,
    ) -> str:
        """
        Erstellt ein Reporting-Sheet auf Basis der übergebenen ReportingRowModel-Daten und speichert es ab.
        """
        template_name: str = template_path / self.config.templates.reporting_template

        try:
            wb = load_workbook(template_path / template_name)
        except Exception as e:
            logger.error(f"Fehler beim Laden des Templates: {e}")
            raise RuntimeError(f"Fehler beim Laden des Templates: {e}")

        ws = wb.active
        if ws is None:
            logger.error("Kein aktives Arbeitsblatt im Template gefunden.")
            raise RuntimeError("Kein aktives Arbeitsblatt im Template gefunden.")

        # Blattschutz deaktivieren, um Felder zu beschreiben
        original_sheet_protected = False
        if hasattr(ws, "protection") and ws.protection and ws.protection.sheet:
            original_sheet_protected = True
            ws.protection.sheet = False

        # Ausfüllen der relevanten Felder im Excel-Sheet
        try:
            ws["c5"] = f"{header_data.employee_first_name} {header_data.employee_last_name}"
            ws["g5"] = header_data.employee_id
            ws["c6"] = reporting_month_dt
            ws["c6"].number_format = "MM.YYYY"
            ws["c7"] = header_data.allowed_hours_per_month
            ws["g7"] = header_data.service_type
            ws["c8"] = header_data.short_code
            ws["g8"] = header_data.client_id
        except Exception as e:
            logger.error(f"Fehler beim Ausfüllen des Sheets: {e}")
            raise RuntimeError(f"Fehler beim Ausfüllen des Sheets: {e}")

        # Blattschutz nur wieder aktivieren, wenn die Originaldatei geschützt war
        if original_sheet_protected:
            ws.protection.sheet = True
            ws.protection.enable()
            password = sheet_password if sheet_password is not None else self.sheet_password
            if password is None:
                logger.error("Sheet-Passwort ist nicht gesetzt!")
                raise RuntimeError("Sheet-Passwort ist nicht gesetzt!")
            ws.protection.set_password(str(password))
            ws.protection.objects = True
            for attr, value in [
                ("enable_select_locked_cells", False),
                ("enable_select_unlocked_cells", True),
                ("format_cells", False),
                ("format_columns", False),
                ("format_rows", False),
                ("insert_columns", False),
                ("insert_rows", False),
                ("insert_hyperlinks", False),
                ("delete_columns", False),
                ("delete_rows", False),
                ("sort", False),
                ("auto_filter", False),
                ("objects", False),
                ("scenarios", False),
            ]:
                if hasattr(ws.protection, attr):
                    setattr(ws.protection, attr, value)

        # Dateinamen generieren und Datei speichern
        dateiname: str = f"{header_data.client_id} ({header_data.short_code})_{reporting_month_dt.strftime('%Y-%m')}.xlsx"
        try:
            wb.save(output_path / dateiname)
        except Exception as e:
            logger.error(f"Fehler beim Speichern der Datei: {e}")
            raise RuntimeError(f"Fehler beim Speichern der Datei: {e}")
        return dateiname


# Beispiel für die Initialisierung mit zentralem Config-Objekt
if __name__ == "__main__":
    config_path = Path(__file__).parent.parent.parent.parent / ".config" / "wegpiraten_config.yaml"
    config = Config(config_path)
    factory = ReportingFactory(config)
    db_path = Path(config.structure.prj_root) / config.structure.local_data_path / config.database.sqlite_db_name
    reporting_month = "2025-10"
    reporting_rows = factory.fetch_reporting_data(db_path, reporting_month)
    if reporting_rows:
        output_path = Path(config.structure.prj_root) / config.structure.output_path
        template_path = Path(config.structure.prj_root) / config.structure.template_path
        new_sheet =factory.create_reporting_sheet(
            header_data=reporting_rows[0],
            reporting_month_dt=datetime.strptime(reporting_month, "%Y-%m"),
            output_path=output_path,
            template_path=template_path,
        )
        logger.info(f"Reporting-Sheet {new_sheet} erfolgreich erstellt.")
