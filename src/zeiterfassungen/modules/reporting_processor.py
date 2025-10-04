from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml  # Wird nur im __main__-Block benötigt, aber für Übersichtlichkeit oben gelassen
from loguru import logger
from openpyxl import load_workbook
from pydantic import ValidationError

from shared_modules.config import Config  # Für Passwort-Handling

from .reporting_config import (
    ReportingConfig,  # Import des Pydantic-Modells für die Reporting-Konfiguration
)
from .reporting_factory import ReportingFactory


class ReportingProcessor:
    """
    Klasse zur Verarbeitung von Reporting-Daten.
    Erwartet ausschließlich Pydantic-Modelle für Konfiguration und nutzt Typsicherheit.
    """

    def get_sheet_password(self) -> str:
        """
        Holt das Excel-Blattschutz-Passwort sicher aus der Umgebung (.env), entschlüsselt falls nötig.
        Gibt SHEET_PASSWORD_ENC (verschlüsselt) oder SHEET_PASSWORD (Klartext) zurück.

        Returns:
            str: Das entschlüsselte oder im Klartext gespeicherte Passwort.

        Raises:
            RuntimeError: Wenn kein Passwort gefunden werden kann.
        """
        config = Config()
        pw = config.get_decrypted_secret("SHEET_PASSWORD_ENC")
        if not pw:
            pw = config.get_secret("SHEET_PASSWORD")
        if not pw:
            logger.error(
                "Excel-Blattschutz-Passwort nicht gesetzt! Bitte .env mit SHEET_PASSWORD_ENC oder SHEET_PASSWORD anlegen."
            )
            raise RuntimeError(
                "Excel-Blattschutz-Passwort nicht gesetzt! Bitte .env mit SHEET_PASSWORD_ENC oder SHEET_PASSWORD anlegen."
            )
        return pw

    def __init__(
        self, reporting_config: ReportingConfig, reporting_factory: ReportingFactory
    ):
        """
        Konstruktor erwartet ein Pydantic-Modell für die Konfiguration.
        Das sorgt für Typsicherheit und Validierung der Konfigurationsdaten.

        Args:
            reporting_config (ReportingConfig): Validierte Reporting-Konfiguration.
            reporting_factory (ReportingFactory): Factory-Objekt zur Erstellung der Reporting-Sheets.
        """
        self.reporting_config: ReportingConfig = reporting_config
        self.reporting_factory = reporting_factory

    def load_client_data(self, reporting_month: str) -> pd.DataFrame:
        """
        Lädt die Klientendaten für den angegebenen Berichtsmonat.
        Nutzt die validierte Pydantic-Konfiguration für alle Pfadangaben.

        Args:
            reporting_month (str): Monat im Format "YYYY-MM".

        Returns:
            pd.DataFrame: Gefilterte Klientendaten.

        Raises:
            ValueError: Wenn die gewünschte Tabelle nicht gefunden wird.
        """
        # Zugriff auf die Konfigurationsdaten über das Pydantic-Modell
        data_path = Path(self.reporting_config.structure.data_path)
        db_name = self.reporting_config.db_name
        source = data_path / db_name
        table_name = self.reporting_config.client_sheet_name

        if not source.exists():
            logger.error(f"Quelldatei für Klientendaten nicht gefunden: {source}")
            raise FileNotFoundError(
                f"Quelldatei für Klientendaten nicht gefunden: {source}"
            )

        reporting_month_dt = datetime.strptime(reporting_month, "%Y-%m")
        wb = load_workbook(source, data_only=True)
        for ws in wb.worksheets:
            if table_name in ws.tables:
                table = ws.tables[table_name]
                ref = table.ref
                data = ws[ref]
                rows = [[cell.value for cell in row] for row in data]
                client_masterdata = pd.DataFrame(rows[1:], columns=rows[0])
                break
        else:
            logger.error(f"Klientendaten {table_name} nicht gefunden in {db_name}")
            raise ValueError(f"Klientendaten {table_name} nicht gefunden in {db_name}")

        # Datumsfilterung: Nur Klienten, deren "Ende" leer ist oder nach dem Berichtsmonat liegt
        client_masterdata["Ende"] = pd.to_datetime(client_masterdata["Ende"], format="%d.%m.%Y", errors="coerce")
        client_masterdata = client_masterdata[(client_masterdata["Ende"].isna()) | (client_masterdata["Ende"] >= reporting_month_dt)]
        return client_masterdata

    def run(self, reporting_month: str, output_path: Path, template_path: Path) -> None:
        """
        Führt die Berichtsverarbeitung für den angegebenen Monat aus.

        Args:
            reporting_month (str): Monat im Format "YYYY-MM".
            output_path (Path): Zielverzeichnis für die erzeugten Dateien.
            template_path (Path): Verzeichnis mit den Excel-Templates.
        """
        reporting_month_dt = datetime.strptime(reporting_month, "%Y-%m")
        client_masterdata = self.load_client_data(reporting_month)
        sheet_password = self.get_sheet_password()
        for _, header_data in client_masterdata.iterrows():
            dateiname = self.reporting_factory.create_reporting_sheet(
                header_data=header_data,
                reporting_month_dt=reporting_month_dt,
                output_path=output_path,
                template_path=template_path,
                sheet_password=sheet_password,
            )
            logger.info(
                f"Erstelle AZ Erfassungsbogen für {header_data['Sozialpädagogin']} "
                f"({header_data['Kürzel']}, Ende: {header_data['Ende']}) -> {dateiname}"
            )


# Beispiel für die Initialisierung mit Pydantic
if __name__ == "__main__":
    # Beispiel: YAML-Konfiguration laden und mit Pydantic validieren
    config_path = Path("wegpiraten_reporting_config.yaml")
    with open(config_path, "r") as f:
        raw_config = yaml.safe_load(f)
    try:
        config = ReportingConfig(**raw_config)
    except ValidationError as e:
        print(f"Konfigurationsfehler: {e}")
        exit(1)

    # Factory-Objekt muss bereitgestellt werden
    factory = None  # Platzhalter
    processor = ReportingProcessor(config, factory)
    # processor.run("2025-08", Path("output"), Path("template.xlsx"))
