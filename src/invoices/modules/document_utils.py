import subprocess
from pathlib import Path
from typing import List, Optional

import pandas as pd
from loguru import logger
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PyPDF2 import PdfMerger

from shared_modules.config import Config

from .invoice_context import InvoiceContext


def _summary_column_letter(summary_df: pd.DataFrame, column_name: str) -> str:
    """Gibt den Excel-Spaltenbuchstaben für eine eindeutige Übersichts-Spalte zurück."""
    loc = summary_df.columns.get_loc(column_name)
    if not isinstance(loc, int):
        raise ValueError(f"Spalte '{column_name}' ist in der Rechnungsübersicht nicht eindeutig.")
    return get_column_letter(loc + 1)


class DocumentUtils:
    """
    Statische Hilfsklasse für Dokumentenoperationen:
    - DOCX nach PDF konvertieren
    - PDFs zusammenführen
    - Rechnungsübersicht als Excel-Tabelle erstellen
    Nutzt konsequent Pydantic-Modelle für Konfiguration und Kontextdaten.
    """

    @staticmethod
    def docx_to_pdf(docx_path: Path, pdf_path: Path, invoice_context: InvoiceContext) -> Path:
        """
        Konvertiert eine DOCX-Datei in eine PDF-Datei mit LibreOffice und benennt sie ggf. nach Vorgabe um.

        Args:
            docx_path (Path): Pfad zur DOCX-Datei.
            pdf_path (Path): Pfad zur Ausgabedatei (PDF).
            invoice_context (InvoiceContext): Kontextobjekt mit allen Rechnungsdaten.
        Returns:
            Path: Pfad zur erzeugten PDF-Datei.
        Raises:
            RuntimeError: Wenn die Konvertierung oder Umbenennung fehlschlägt.
        """
        try:
            subprocess.run(
                [
                    "libreoffice",
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    str(pdf_path.parent),
                    str(docx_path),
                ],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                check=True,
            )
        except Exception as e:
            logger.error(f"PDF-Konvertierung fehlgeschlagen: {e}")
            raise RuntimeError(f"PDF-Konvertierung fehlgeschlagen: {e}")

        _ = invoice_context
        generated_pdf = docx_path.with_suffix(".pdf")
        target_pdf = pdf_path
        try:
            if generated_pdf.resolve() != target_pdf.resolve():
                generated_pdf.replace(target_pdf)
        except Exception as e:
            logger.error(f"PDF konnte nicht umbenannt werden: {e}")
            raise RuntimeError(f"PDF konnte nicht umbenannt werden: {e}")
        logger.debug(f"{target_pdf.name} erzeugt")
        return target_pdf

    @staticmethod
    def merge_pdfs(pdf_files: List[Path], payer_context: InvoiceContext, output_path: Optional[Path] = None) -> Path:
        """
        Führt mehrere PDF-Dateien zu einer einzigen zusammen.

        Args:
            pdf_files (List[Path]): Liste der PDF-Dateipfade.
            payer_context (InvoiceContext): Kontextobjekt mit Zahlungsdienstleister und Zeitraum.
            output_path (Path, optional): Zielverzeichnis für die Sammel-PDF.
        Returns:
            Path: Pfad zur erzeugten PDF-Datei.
        Raises:
            ValueError: Wenn keine PDF-Dateien übergeben wurden.
            FileNotFoundError: Wenn eine der PDF-Dateien nicht gefunden wird.
            RuntimeError: Wenn die Zusammenführung fehlschlägt.
        """
        if not pdf_files:
            raise ValueError("Keine PDF-Dateien zum Zusammenführen gefunden.")
        merger = PdfMerger()
        for pdf_file in pdf_files:
            if not pdf_file.exists():
                raise FileNotFoundError(f"PDF-Datei nicht gefunden: {pdf_file}")
            merger.append(pdf_file)
        # Zielverzeichnis bestimmen
        if output_path is None:
            output_path = pdf_files[0].parent
        payer = payer_context.data.get("payer")
        invoice_month = payer_context.data.get("invoice_month", "unbekannt")
        payer_id = getattr(payer, "key", "n.a") if payer else "n.a"
        merged_pdf_path = output_path / f"Rechnungen_{payer_id}_{invoice_month}.pdf"
        try:
            merger.write(merged_pdf_path)
            merger.close()
        except Exception as e:
            logger.error(f"PDF-Zusammenführung fehlgeschlagen: {e}")
            raise RuntimeError(f"PDF-Zusammenführung fehlgeschlagen: {e}")
        return merged_pdf_path

    @staticmethod
    def create_summary(
        config: Config,
        invoice_list: List[InvoiceContext],
    ) -> Path:
        """
        Erstellt eine Excel-Datei mit der Rechnungsübersicht.
        Es werden nur die Felder invoice_date, payer_id, client_id, invoice_month und Summe_Kosten aufgenommen.
        Args:
            config (Config): Konfigurationsobjekt für Pfade und Formate.
            invoice_list (List[InvoiceContext]): Liste der Rechnungskontexte.
        Returns:
            Path: Pfad zur erzeugten Excel-Datei.
        Raises:
            RuntimeError: Wenn die Datei nicht geschrieben werden kann.
        """
        # Zugriff auf die Pydantic-basierte Konfiguration
        output_path: Path = config.get_output_path()
        kosten_format: str = config.formatting.currency_format or "#,##0.00 ¤"
        datum_format: str = config.formatting.date_format or "dd.MM.yyyy"

        # Nur die gewünschten Felder extrahieren
        summary_rows = []
        for invoice in invoice_list:
            data = invoice.as_dict()
            payer = data.get("payer")
            client = data.get("client")
            soll_f = int(data.get("allowed_travel_time") or 0)
            soll_d = int(data.get("allowed_direct_effort") or 0)
            soll_i = int(data.get("allowed_indirect_effort") or 0)
            ist_f = int(data.get("summe_fahrtzeit") or 0)
            ist_d = int(data.get("summe_direkt") or 0)
            ist_i = int(data.get("summe_indirekt") or 0)
            summary_rows.append(
                {
                    "Rechnungsdatum": data.get("invoice_date"),
                    "Zahlungsträger": getattr(payer, "name", "n.a") if payer else "n.a",
                    "Klienten-ID": getattr(client, "key", "") if client else "",
                    "Klient": (
                        ", ".join(p for p in [getattr(client, "last_name", ""), getattr(client, "first_name", "")] if p)
                        if client
                        else "n.a"
                    ),
                    "Büro": data.get("tenant_city", ""),
                    "Abrechnungsmonat": data.get("invoice_month", "unbekannt"),
                    "Max F": soll_f,
                    "Ist F": ist_f,
                    "Max D": soll_d,
                    "Ist D": ist_d,
                    "Max I": soll_i,
                    "Ist I": ist_i,
                    "Max Total": soll_f + soll_d + soll_i,
                    "Summe Ist": int(data.get("summe_stunden") or 0),
                    "Rechnungsbetrag": data.get("summe_kosten", -999),
                }
            )
        summary_df = pd.DataFrame(summary_rows)
        # Prüfe, ob alle Rechnungen denselben Abrechnungsmonat haben und propagiere diesen
        invoice_months = set(row["Abrechnungsmonat"] for row in summary_rows)
        if len(invoice_months) == 1:
            invoice_month = invoice_months.pop()
        else:
            logger.warning(f"Mehrere Abrechnungsmonate in der Übersicht: {invoice_months}")
            invoice_month = "gemischt"

        summary_file = output_path / f"Rechnungsuebersicht_{invoice_month}.xlsx"
        try:
            with pd.ExcelWriter(summary_file, engine="openpyxl") as writer:
                summary_df.to_excel(writer, index=False, sheet_name="Rechnungsübersicht")
                worksheet = writer.sheets["Rechnungsübersicht"]
                end_col_idx = len(summary_df.columns)
                end_col_letter = get_column_letter(end_col_idx)
                table_ref = f"A1:{end_col_letter}{len(summary_df) + 1}"
                tab = Table(displayName="Rechnungsübersicht", ref=table_ref)
                tab.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                worksheet.add_table(tab)
                total_row_idx = len(summary_df) + 2
                worksheet[f"A{total_row_idx}"] = "Gesamt"
                bold_font = Font(bold=True)
                for col in range(1, end_col_idx + 1):
                    worksheet.cell(row=total_row_idx, column=col).font = bold_font

                # Währungsspalte formatieren und summieren
                if "Rechnungsbetrag" in summary_df.columns:
                    kosten_col_letter = _summary_column_letter(summary_df, "Rechnungsbetrag")
                    for row in range(2, len(summary_df) + 2):
                        worksheet[f"{kosten_col_letter}{row}"].number_format = kosten_format
                    worksheet[f"{kosten_col_letter}{total_row_idx}"] = (
                        f"=SUM({kosten_col_letter}2:{kosten_col_letter}{total_row_idx - 1})"
                    )
                    worksheet[f"{kosten_col_letter}{total_row_idx}"].number_format = kosten_format

                # Integer-Minuten-Spalten formatieren und summieren
                int_min_cols = [
                    "Max F",
                    "Ist F",
                    "Max D",
                    "Ist D",
                    "Max I",
                    "Ist I",
                    "Max Total",
                    "Summe Ist",
                ]
                for col_name in int_min_cols:
                    if col_name in summary_df.columns:
                        col_letter = _summary_column_letter(summary_df, col_name)
                        for row in range(2, len(summary_df) + 2):
                            worksheet[f"{col_letter}{row}"].number_format = "0"
                        worksheet[f"{col_letter}{total_row_idx}"] = (
                            f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})"
                        )
                        worksheet[f"{col_letter}{total_row_idx}"].number_format = "0"

                # Bedingte Formatierung: hellrot wenn Ist > Max
                red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                data_row_count = len(summary_df)
                if data_row_count > 0:
                    for max_col, ist_col in [
                        ("Max F", "Ist F"),
                        ("Max D", "Ist D"),
                        ("Max I", "Ist I"),
                        ("Max Total", "Summe Ist"),
                    ]:
                        if max_col in summary_df.columns and ist_col in summary_df.columns:
                            max_letter = _summary_column_letter(summary_df, max_col)
                            ist_letter = _summary_column_letter(summary_df, ist_col)
                            cell_range = f"{max_letter}2:{ist_letter}{data_row_count + 1}"
                            rule = FormulaRule(formula=[f"${ist_letter}2>${max_letter}2"], fill=red_fill)
                            worksheet.conditional_formatting.add(cell_range, rule)

                # Datumsspalte formatieren
                if "Rechnungsdatum" in summary_df.columns:
                    datum_col_letter = _summary_column_letter(summary_df, "Rechnungsdatum")
                    for row in range(2, len(summary_df) + 2):
                        worksheet[f"{datum_col_letter}{row}"].number_format = datum_format
        except Exception as e:
            logger.error(f"Fehler beim Schreiben der Excel-Datei: {e}")
            raise RuntimeError(f"Fehler beim Schreiben der Excel-Datei: {e}")
        return summary_file


if __name__ == "__main__":
    print("DocumentUtils Modul. Nicht direkt ausführbar.")
