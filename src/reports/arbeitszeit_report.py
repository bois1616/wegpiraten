"""
Erstellt ein Arbeitszeitprotokoll als Excel-Datei aus service_data.
Zeiten werden pro Mitarbeiter, Klient und Datum aufgelistet und pro Mitarbeiter summiert.
"""

import sqlite3
from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from shared_modules.config import Config
from shared_modules.month_period import get_month_period
from shared_modules.utils import ensure_dir

_SQL = """
SELECT
    e.last_name || ' ' || e.first_name        AS mitarbeiter,
    sd.employee_id,
    c.last_name || ' ' || c.first_name        AS klient,
    sd.client_id,
    sd.service_date,
    COALESCE(sd.travel_time, 0)               AS fahrtzeit,
    COALESCE(sd.direct_time, 0)               AS direkt,
    COALESCE(sd.indirect_time, 0)             AS indirekt,
    COALESCE(sd.travel_time, 0)
        + COALESCE(sd.direct_time, 0)
        + COALESCE(sd.indirect_time, 0)       AS total
FROM service_data sd
LEFT JOIN employees e  ON sd.employee_id = e.emp_id
LEFT JOIN clients   c  ON sd.client_id   = c.client_id
WHERE sd.reporting_month = ?
ORDER BY e.last_name, e.first_name, sd.service_date, c.last_name
"""

_HEADER = [
    "Mitarbeiter",
    "MA-ID",
    "Klient",
    "Klienten-ID",
    "Datum",
    "Fahrtzeit",
    "Direktkontakt",
    "Indir. Bearbeitung",
    "Total",
]

_TIME_COLS = ["Fahrtzeit", "Direktkontakt", "Indir. Bearbeitung", "Total"]


def create_arbeitszeit_report(config: Config, reporting_month: str) -> Path:
    """
    Erstellt ein Arbeitszeitprotokoll für den angegebenen Monat.

    Args:
        config: Konfigurationsobjekt.
        reporting_month: Abrechnungsmonat (beliebiges Format MM.YYYY / YYYY-MM).

    Returns:
        Pfad zur erzeugten Excel-Datei.
    """
    period = get_month_period(reporting_month)
    month_str = period.start.strftime("%Y-%m")
    output_path = ensure_dir(config.get_output_path())
    out_file = output_path / f"Arbeitszeitprotokoll_{month_str}.xlsx"

    db_path = config.get_db_path()
    with sqlite3.connect(db_path) as conn:
        df = pd.read_sql_query(_SQL, conn, params=(month_str,))

    if df.empty:
        logger.warning("Keine Leistungsdaten für {} in service_data.", month_str)
        raise ValueError(f"Keine Leistungsdaten für {month_str} gefunden.")

    df.columns = _HEADER  # type: ignore[assignment]
    df["Datum"] = pd.to_datetime(df["Datum"]).dt.date

    _write_excel(df, out_file, month_str)
    logger.info("Arbeitszeitprotokoll geschrieben: {}", out_file)
    return out_file


def _write_excel(df: pd.DataFrame, out_file: Path, month_str: str) -> None:
    employees = df["Mitarbeiter"].unique()

    # Summenzeilen pro Mitarbeiter
    summary_rows = []
    for emp in employees:
        emp_df = df[df["Mitarbeiter"] == emp]
        summary_rows.append(
            {
                "Mitarbeiter": emp,
                "MA-ID": emp_df["MA-ID"].iloc[0],
                "Klient": "",
                "Klienten-ID": "",
                "Datum": None,
                "Fahrtzeit": emp_df["Fahrtzeit"].sum(),
                "Direktkontakt": emp_df["Direktkontakt"].sum(),
                "Indir. Bearbeitung": emp_df["Indir. Bearbeitung"].sum(),
                "Total": emp_df["Total"].sum(),
            }
        )
    summary_df = pd.DataFrame(summary_rows)

    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subtotal_font = Font(bold=True)
    date_fmt = "dd.mm.yyyy"
    int_fmt = "0"

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        # --- Detailtabelle ---
        df.to_excel(writer, sheet_name="Protokoll", index=False)
        ws = writer.sheets["Protokoll"]
        _apply_table(ws, df, "Protokoll")
        _format_sheet(ws, df, header_fill, header_font, date_fmt, int_fmt)

        # --- Zusammenfassung ---
        summary_df.to_excel(writer, sheet_name="Zusammenfassung", index=False)
        ws_sum = writer.sheets["Zusammenfassung"]
        _apply_table(ws_sum, summary_df, "Zusammenfassung")
        _format_sheet(ws_sum, summary_df, header_fill, header_font, date_fmt, int_fmt)

        # Subtotal-Zeilen hervorheben
        for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=2):
            for col_idx in range(1, len(summary_df.columns) + 1):
                cell = ws_sum.cell(row=row_idx, column=col_idx)
                cell.fill = subtotal_fill
                cell.font = subtotal_font

        # Gesamtzeile in Zusammenfassung
        total_row = len(summary_df) + 2
        ws_sum.cell(row=total_row, column=1, value="Gesamt").font = Font(bold=True)
        for col_name in _TIME_COLS:
            col_idx = list(summary_df.columns).index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            ws_sum.cell(
                row=total_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
            ).font = Font(bold=True)
            ws_sum.cell(row=total_row, column=col_idx).number_format = int_fmt

        # --- Pivot: Zeitdaten je Mitarbeiter und Tag ---
        pivot_df = _build_pivot(df)
        pivot_df.to_excel(writer, sheet_name="Pivot", index=False)
        ws_pivot = writer.sheets["Pivot"]
        _write_pivot_sheet(ws_pivot, pivot_df, header_fill, header_font)


def _apply_table(ws, df: pd.DataFrame, name: str) -> None:
    end_col = get_column_letter(len(df.columns))
    ref = f"A1:{end_col}{len(df) + 1}"
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _format_sheet(ws, df: pd.DataFrame, header_fill, header_font, date_fmt: str, int_fmt: str) -> None:
    cols = list(df.columns)
    last_data_row = len(df) + 1

    # Kopfzeile
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datumsspalte
    if "Datum" in cols:
        date_col = get_column_letter(cols.index("Datum") + 1)
        for row in range(2, last_data_row + 1):
            ws[f"{date_col}{row}"].number_format = date_fmt
            ws[f"{date_col}{row}"].alignment = Alignment(horizontal="center")

    # Zeitspalten
    for col_name in _TIME_COLS:
        if col_name not in cols:
            continue
        col_letter = get_column_letter(cols.index(col_name) + 1)
        for row in range(2, last_data_row + 1):
            ws[f"{col_letter}{row}"].number_format = int_fmt
            ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="right")

    # Spaltenbreiten anpassen
    _autofit(ws, df)


def _autofit(ws, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if not df.empty else 0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)


_PIVOT_MIN_COLS = ["Sum Fahrzeit", "Sum Direkt", "Sum Indirekt", "Gesamt Min"]
_PIVOT_HMM_COL = "Gesamt h:mm"


def _build_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregiert Zeitdaten je Mitarbeiter und Tag."""
    pivot = (
        df.groupby(["Mitarbeiter", "Datum"], as_index=False)
        .agg(
            **{
                "Sum Fahrzeit": ("Fahrtzeit", "sum"),
                "Sum Direkt": ("Direktkontakt", "sum"),
                "Sum Indirekt": ("Indir. Bearbeitung", "sum"),
            }
        )
        .sort_values(["Mitarbeiter", "Datum"])
    )
    pivot["Gesamt Min"] = pivot["Sum Fahrzeit"] + pivot["Sum Direkt"] + pivot["Sum Indirekt"]
    # Platzhalter; die eigentliche h:mm-Formel wird beim Schreiben als Excel-Formel eingetragen
    pivot[_PIVOT_HMM_COL] = None
    return pivot  # type: ignore[return-value]


def _write_pivot_sheet(ws, pivot_df: pd.DataFrame, header_fill, header_font) -> None:
    """Formatiert das Pivot-Sheet und setzt die [h]:mm-Formel für Gesamt h:mm."""
    from openpyxl.styles import Alignment as Aln

    cols = list(pivot_df.columns)
    min_col_idx = cols.index("Gesamt Min") + 1
    hmm_col_idx = cols.index(_PIVOT_HMM_COL) + 1
    min_col_letter = get_column_letter(min_col_idx)
    date_col_idx = cols.index("Datum") + 1
    date_col_letter = get_column_letter(date_col_idx)

    # Kopfzeile formatieren
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Aln(horizontal="center")

    last_row = len(pivot_df) + 1
    for row in range(2, last_row + 1):
        # Datumsspalte
        ws[f"{date_col_letter}{row}"].number_format = "dd.mm.yyyy"
        ws[f"{date_col_letter}{row}"].alignment = Aln(horizontal="center")
        # Minuten-Spalten
        for col_name in _PIVOT_MIN_COLS:
            col_idx = cols.index(col_name) + 1
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = "0"
            cell.alignment = Aln(horizontal="right")
        # h:mm-Formel: Minuten / 1440 ergibt einen Excel-Zeitwert; Format [h]:mm zeigt >24h korrekt
        hmm_cell = ws.cell(row=row, column=hmm_col_idx, value=f"={min_col_letter}{row}/1440")
        hmm_cell.number_format = "[h]:mm"
        hmm_cell.alignment = Aln(horizontal="right")

    # Spaltenbreiten
    for col_idx, col_name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(len(str(col_name)) + 4, 22)
